#!/usr/bin/env python3
"""
script_protocol_generator.py — FAT/SAT Test Protocol Generator (Phase 27-A;
SAT v2 Faz 1)

Produces a FAT or SAT test protocol from the project metadata (RD01, RD08,
PROJECT_STATE) in Markdown, optional Excel and optional PDF.

FAT sections: Pre-Checks (workshop), IO Test (PLCSim), Function Test, Alarm
Test.  SAT sections: Pre-Checks (site: field devices, as-built docs), Loop
Check with real field devices, Function Test, Alarm Test.  The full SAT site
acceptance catalogue (motor rotation, real E-stop chain, drive parameters,
cybersecurity, backup/handover) lives in fat_protocol.run_fat_protocol —
this CLI generator is the RD-table-driven variant.

Every row carries a "Ref." column naming the standard it derives from
(IEC 62381 / IEC 62061 / ISO 13849-2 / EN 60204-1 / IEC 62682).
Languages: DE (default) / EN / TR via protocol_i18n.

CLI:
  python script_protocol_generator.py --project PROJECT_PATH
      [--type FAT|SAT|BOTH] [--lang de|en|tr] [--pdf] [--excel]
"""

from __future__ import annotations

import re
import json
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

# B-P5 / S-17: RD05 readiness gate — shared with fat_protocol.py.
# Import the single source-of-truth helper; do NOT duplicate the logic here.
try:
    import sys as _sys
    _SCRIPTS = Path(__file__).resolve().parent
    if str(_SCRIPTS) not in _sys.path:
        _sys.path.insert(0, str(_SCRIPTS))
    from fat_protocol import check_rd05_ready, Rd05BlockedError
except ImportError:  # pragma: no cover — fat_protocol.py must always be present
    def check_rd05_ready(project_path):  # type: ignore[misc]
        raise RuntimeError("fat_protocol.check_rd05_ready could not be imported")
    class Rd05BlockedError(RuntimeError):  # type: ignore[misc]
        pass

from protocol_i18n import (  # noqa: E402
    DEFAULT_LANG, force_utf8_stdout, normalize_lang, t,
)

# R-S-1: merkezi F-CPU tespit yardımcısı
try:
    from workbench.core.safety_utils import is_f_cpu as _is_f_cpu
except ImportError:  # pragma: no cover
    def _is_f_cpu(cpu_model):  # type: ignore[misc]
        import re as _re
        return bool(_re.search(r"(?i)(\bSF\b|^SF|\bTF\b|\dF[-\s/]|\dF$|\dTF[-\s/]|\dTF$|[-\s]F[-\s]|[-\s]F$)", cpu_model or ""))


FACTORY_ROOT = Path(__file__).resolve().parent.parent


# -- Data structures ----------------------------------------------------------

@dataclass
class TestItem:
    section: str          # "IO" / "Alarm" / "Function" / "PreCheck"
    id: str               # T-IO-001 etc.
    description: str
    expected: str
    tag: str = ""
    ref: str = ""         # standard the row derives from (Ref. column)
    result_placeholder: str = "[ ]"  # [ ] pending / [x] passed / [!] failed


@dataclass
class ProtocolResult:
    project_name: str
    test_type: str        # "FAT" / "SAT"
    lang: str = DEFAULT_LANG
    items: list[TestItem] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    # raw RD08 alarm dicts incl. IEC 62682 fields (priority/response/class)
    alarms_meta: list[dict] = field(default_factory=list)
    md_path: Optional[Path] = None
    xlsx_path: Optional[Path] = None
    pdf_path: Optional[Path] = None

    @property
    def io_items(self):
        return [i for i in self.items if i.section == "IO"]

    @property
    def alarm_items(self):
        return [i for i in self.items if i.section == "Alarm"]

    @property
    def function_items(self):
        return [i for i in self.items if i.section == "Function"]

    @property
    def precheck_items(self):
        return [i for i in self.items if i.section == "PreCheck"]


# -- Source readers -----------------------------------------------------------

def _read_state(project_path: Path) -> dict:
    f = project_path / "PROJECT_STATE.json"
    if f.exists():
        try:
            return json.loads(f.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def _read_rd(project_path: Path, rd_id: str) -> str:
    """Read an RD file from the project's metadata/ folder."""
    for f in (project_path / "metadata").glob(f"{rd_id}*.md"):
        try:
            return f.read_text(encoding="utf-8", errors="replace")
        except Exception:
            pass
    return ""


def _scan_scl_blocks(project_path: Path) -> list[dict]:
    """Return the FB/FC names in the _output/scl/ folder."""
    scl_dir = project_path / "_output" / "scl"
    blocks: list[dict] = []
    if not scl_dir.exists():
        return blocks
    fb_re = re.compile(r"FUNCTION_BLOCK\s+[\"']?(\w+)[\"']?", re.IGNORECASE)
    fc_re = re.compile(r"^FUNCTION\s+[\"']?(\w+)[\"']?", re.IGNORECASE | re.MULTILINE)
    for scl_file in sorted(scl_dir.glob("*.scl")):
        try:
            content = scl_file.read_text(encoding="utf-8", errors="replace")
        except Exception:
            continue
        for m in fb_re.finditer(content):
            blocks.append({"type": "FB", "name": m.group(1)})
            break
        for m in fc_re.finditer(content):
            blocks.append({"type": "FC", "name": m.group(1)})
            break
    return blocks


def _split_row(stripped: str) -> list[str]:
    """Split a markdown table row into cells with stable indices.

    Outer pipes are removed BEFORE splitting; inner empty cells are kept.
    The previous filter dropped only leading/trailing empties from the
    header but kept them in data rows, shifting every column by one — rows
    were then silently discarded (name == "").
    """
    return [c.strip() for c in stripped.strip("|").split("|")]


def _parse_io_from_rd01(rd01_text: str) -> list[dict]:
    """Extract IO signals from the markdown table in RD01.
    (Turkish column keywords kept to match legacy/Turkish RD01 data.)"""
    signals: list[dict] = []
    in_table = False
    col_names: list[str] = []
    for line in rd01_text.splitlines():
        stripped = line.strip()
        if stripped.startswith("|") and not in_table:
            lower = [c.lower() for c in _split_row(stripped)]
            if any(k in " ".join(lower) for k in ("tag", "sinyal", "signal", "giriş", "çıkış")):
                col_names = lower
                in_table = True
                continue
        if in_table and re.match(r"^\|[-|\s:]+\|$", stripped):
            continue
        if in_table and stripped.startswith("|"):
            parts = _split_row(stripped)
            if not any(parts):
                in_table = False
                continue

            def _col(keys):
                for k in keys:
                    for i, cn in enumerate(col_names):
                        if k in cn and i < len(parts):
                            return parts[i]
                return parts[0] if parts else ""

            name = _col(["tag", "sinyal", "signal", "ad", "name"])
            addr = _col(["adres", "address", "%i", "%q"])
            tip  = _col(["tip", "type", "yön", "direction"])
            if name and name != "-":
                signals.append({"name": name, "address": addr, "type": tip})
        elif in_table and not stripped.startswith("|"):
            in_table = False
    return signals


def _parse_alarms_from_rd08(rd08_text: str) -> list[dict]:
    """Extract the alarm list from the markdown table in RD08.

    Faz 5 (IEC 62682): besides message/tag/priority the parser now also
    looks for operator-response and alarm-class columns.  Missing values
    stay empty here; the writer renders them as "—" plus a fill-in
    instruction (silent AI guessing is forbidden — values come from RD08
    or from the engineer, never invented).
    (Turkish column keywords kept to match legacy/Turkish RD08 data.)
    """
    alarms: list[dict] = []
    in_table = False
    col_names: list[str] = []
    for line in rd08_text.splitlines():
        stripped = line.strip()
        if stripped.startswith("|") and not in_table:
            lower = [c.lower() for c in _split_row(stripped)]
            if any(k in " ".join(lower) for k in ("alarm", "hata", "fault", "message", "mesaj")):
                col_names = lower
                in_table = True
                continue
        if in_table and re.match(r"^\|[-|\s:]+\|$", stripped):
            continue
        if in_table and stripped.startswith("|"):
            parts = _split_row(stripped)
            if not any(parts):
                in_table = False
                continue

            def _col(keys, dflt=""):
                for k in keys:
                    for i, cn in enumerate(col_names):
                        if k in cn and i < len(parts):
                            return parts[i]
                return dflt

            def _col_or_first(keys):
                v = _col(keys)
                return v if v else (parts[0] if parts else "")

            msg  = _col_or_first(["alarm", "mesaj", "message", "text", "açıklama", "description"])
            tag  = _col(["tag", "sinyal", "signal", "bit"])
            prio = _col(["öncelik", "priority", "sever", "prio"])
            resp = _col(["tepki", "response", "reaktion", "operator", "operatör"])
            acls = _col(["sınıf", "class", "klasse", "kategorie", "kategori"])
            if msg and msg != "-":
                alarms.append({
                    "message": msg, "tag": tag, "priority": prio,
                    "response": resp, "alarm_class": acls,
                })
        elif in_table and not stripped.startswith("|"):
            in_table = False
    return alarms


# -- Test item builders -------------------------------------------------------

_COUNTER: dict[str, int] = {}

def _next_id(section_code: str) -> str:
    _COUNTER[section_code] = _COUNTER.get(section_code, 0) + 1
    return f"T-{section_code}-{_COUNTER[section_code]:03d}"


def _build_prechecks(state: dict, lang: str = DEFAULT_LANG,
                     test_type: str = "FAT") -> list[TestItem]:
    items: list[TestItem] = []
    platform = state.get("target_platform", "")
    cpu = state.get("target_cpu", "")
    tia_ver = state.get("target_tia_version", "")
    ref_62381 = t("ref.iec62381", lang)
    ref_60204 = t("ref.en60204_1", lang)

    def pre(key: str, ref: str, **fmt) -> TestItem:
        return TestItem("PreCheck", _next_id("PRE"),
                        t(key, lang, **fmt), t(key + ".expected", lang, **fmt),
                        ref=ref)

    # Hardware
    items.append(pre("pre.cpu_power", ref_60204))
    items.append(pre("pre.cables", ref_60204))
    if "PROFINET" in (platform + cpu).upper() or state.get("selected_devices"):
        items.append(pre("pre.profinet", ref_62381))
    if test_type == "SAT":
        items.append(pre("pre.sat_field_devices", ref_62381))
        items.append(pre("pre.sat_asbuilt", ref_62381))
    # TIA Portal
    items.append(pre("pre.tia_compile", ref_62381, ver=tia_ver or "V18+"))
    items.append(pre("pre.download", ref_62381))
    items.append(pre("pre.run_mode", ref_62381))
    # Safety (if any)
    if _is_f_cpu(cpu):  # R-S-1: merkezi F-CPU tespiti
        items.append(pre("pre.fcpu_check",
                         f"{t('ref.iec62061', lang)} / {t('ref.iso13849_2', lang)}"))
    return items


def _build_io_tests(signals: list[dict], lang: str, test_type: str) -> list[TestItem]:
    """FAT: PLCSim IO test wording.  SAT: loop check with real field devices."""
    items: list[TestItem] = []
    ref = t("ref.iec62381", lang)
    for sig in signals:
        name = sig.get("name", "")
        addr = sig.get("address", "")
        sig_type = sig.get("type", "").upper()
        if not name:
            continue
        is_input = any(k in sig_type for k in ("DI", "AI", "INPUT", "GİRİŞ", "GIRISH"))
        is_output = any(k in sig_type for k in ("DQ", "DO", "AO", "AQ", "OUTPUT", "ÇIKIŞ"))
        is_analog = any(k in sig_type for k in ("AI", "AO", "AQ", "ANALOG"))
        shown_addr = addr or name
        if test_type == "SAT":
            if is_analog and is_input:
                desc_key, exp_key = "sat.loop.row_ai", "sat.loop.expected_ai"
            elif is_analog and is_output:
                desc_key, exp_key = "sat.loop.row_ao", "sat.loop.expected_ao"
            elif is_output:
                desc_key, exp_key = "sat.loop.row_do", "sat.loop.expected_do"
            else:
                desc_key, exp_key = "sat.loop.row_di", "sat.loop.expected_di"
            desc = t(desc_key, lang, name=name)
            expected = t(exp_key, lang, addr=shown_addr)
        else:
            if is_input:
                desc = t("io.test_input", lang, name=name)
                expected = t("io.test_input.expected", lang, addr=shown_addr)
            elif is_output:
                desc = t("io.test_output", lang, name=name)
                expected = t("io.test_output.expected", lang, addr=shown_addr)
            else:
                desc = t("io.test_generic", lang, name=name)
                expected = t("io.test_generic.expected", lang, addr=shown_addr)
        items.append(TestItem("IO", _next_id("IO"), desc, expected, tag=name, ref=ref))
    return items


def _build_function_tests(blocks: list[dict], lang: str) -> list[TestItem]:
    items: list[TestItem] = []
    ref = t("ref.iec62381", lang)
    for blk in blocks:
        btype = blk["type"]
        bname = blk["name"]
        items.append(TestItem(
            "Function", _next_id("FCT"),
            t("func.normal", lang, btype=btype, bname=bname),
            t("func.normal.expected", lang, bname=bname),
            tag=bname, ref=ref,
        ))
        items.append(TestItem(
            "Function", _next_id("FCT"),
            t("func.fault", lang, btype=btype, bname=bname),
            t("func.fault.expected", lang, bname=bname),
            tag=bname, ref=ref,
        ))
    return items


def _build_alarm_tests(alarms: list[dict], lang: str) -> list[TestItem]:
    items: list[TestItem] = []
    ref = t("ref.iec62682", lang)
    for alarm in alarms:
        msg = alarm.get("message", "")
        tag = alarm.get("tag", "")
        item = TestItem(
            "Alarm", _next_id("ALM"),
            t("alarm.test", lang, msg=msg),
            t("alarm.test.expected", lang, msg=msg),
            tag=tag, ref=ref,
        )
        items.append(item)
    return items


# -- Protocol generator -------------------------------------------------------

def generate_protocol(
    project_path: Path,
    test_type: str = "FAT",
    lang: str = DEFAULT_LANG,
) -> ProtocolResult:
    """Generate a full FAT or SAT test protocol (single type — BOTH is
    handled by ``run_protocol_set``).

    B-P5 / S-17: Raises ``Rd05BlockedError`` when RD05 (Safety Requirements)
    is absent, empty, or template-only.  No protocol file is written in that
    case.  Fail-closed: when in doubt, block.
    """
    test_type = (test_type or "FAT").upper()
    if test_type not in ("FAT", "SAT"):
        raise ValueError(
            f"generate_protocol: test_type must be FAT or SAT (got {test_type!r})"
        )
    lang = normalize_lang(lang)

    # RD05 gate — must be first; propagates Rd05BlockedError to caller.
    check_rd05_ready(project_path)

    _COUNTER.clear()
    state = _read_state(project_path)
    rd01  = _read_rd(project_path, "RD01")
    rd08  = _read_rd(project_path, "RD08")
    blocks = _scan_scl_blocks(project_path)

    result = ProtocolResult(
        project_name=project_path.name,
        test_type=test_type,
        lang=lang,
    )

    result.items += _build_prechecks(state, lang, test_type)
    signals = _parse_io_from_rd01(rd01)
    if signals:
        result.items += _build_io_tests(signals, lang, test_type)
    else:
        result.warnings.append(t("io.no_table_warning", lang))

    result.items += _build_function_tests(blocks, lang)

    alarms = _parse_alarms_from_rd08(rd08)
    if alarms:
        result.items += _build_alarm_tests(alarms, lang)
        result.alarms_meta = alarms
    else:
        result.warnings.append(t("alarm.no_table_warning", lang))

    return result


def write_protocol_md(result: ProtocolResult, project_path: Path) -> Path:
    """Write the protocol to a Markdown file."""
    lang = result.lang
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    ts_file = datetime.now().strftime("%Y%m%d_%H%M")
    title = t("common.sat_title" if result.test_type == "SAT" else "common.fat_title", lang)
    lines: list[str] = [
        f"# {title} — {result.project_name}",
        "",
        "```yaml",
        f"test_type: {result.test_type}",
        f"lang: {lang}",
        f"project: {result.project_name}",
        f"created: {ts}",
        f"total_tests: {len(result.items)}",
        "```",
        "",
        f"> {t('common.auto_generated', lang)}",
        f"> {t('common.engineer_additions', lang)}",
        "",
    ]

    def _section(title: str, items: list[TestItem]):
        if not items:
            return
        lines.append(f"## {title}")
        lines.append("")
        lines.append(
            f"| {t('col.test_id', lang)} | {t('col.description', lang)} "
            f"| {t('col.tag', lang)} | {t('col.expected', lang)} "
            f"| {t('col.ref', lang)} | {t('col.actual', lang)} "
            f"| {t('col.date', lang)} | {t('col.signature', lang)} |"
        )
        lines.append("|---|---|---|---|---|---|---|---|")
        for item in items:
            lines.append(
                f"| {item.id} | {item.description} | `{item.tag or '-'}` "
                f"| {item.expected} | {item.ref or '-'} "
                f"| {item.result_placeholder} | ______ | ______ |"
            )
        lines.append("")

    io_title = t("sec.sat_loop_check" if result.test_type == "SAT" else "sec.io_test", lang)
    _section(f"1. {t('sec.precheck', lang)}", result.precheck_items)
    _section(f"2. {io_title}", result.io_items)
    _section(f"3. {t('sec.function_test', lang)}", result.function_items)
    _section(f"4. {t('sec.alarm_test', lang)}", result.alarm_items)

    # Faz 5 — IEC 62682 rationalization table (priority / operator response /
    # alarm class).  Values come from RD08; missing ones are rendered as "—"
    # with a fill-in instruction.  AI guessing is forbidden.
    alarms_meta = result.alarms_meta
    if alarms_meta:
        missing_any = any(
            not (a.get("priority") and a.get("response") and a.get("alarm_class"))
            for a in alarms_meta
        )
        lines.append(f"### {t('alarm62682.title', lang)}")
        lines.append("")
        lines.append(f"> {t('alarm62682.intro', lang)}")
        if missing_any:
            lines.append(f"> ⚠ {t('alarm62682.fill_instruction', lang)}")
        lines.append("")
        lines.append(
            f"| {t('col.no', lang)} | {t('col.description', lang)} "
            f"| {t('alarm62682.col_priority', lang)} "
            f"| {t('alarm62682.col_response', lang)} "
            f"| {t('alarm62682.col_class', lang)} | {t('col.ref', lang)} |"
        )
        lines.append("|---|---|---|---|---|---|")
        ref62682 = t("ref.iec62682", lang)
        for idx, a in enumerate(alarms_meta, start=1):
            lines.append(
                f"| {idx} | {a.get('message', '')} "
                f"| {a.get('priority') or '—'} | {a.get('response') or '—'} "
                f"| {a.get('alarm_class') or '—'} | {ref62682} |"
            )
        lines.append("")

    lines += [
        f"## 5. {t('common.general_observations', lang)}",
        "",
        f"> {t('common.general_observations_hint', lang)}",
        "",
        f"## 6. {t('common.signatures', lang)}",
        "",
        "| | |",
        "|---|---|",
        f"| {t('common.test_lead', lang)} | ________________________ |",
        f"| {t('common.customer_rep', lang)} | ________________________ |",
        f"| {t('common.date', lang)} | ________________________ |",
        f"| {t('common.test_result', lang)} | [ ] {t('status.pending', lang)} |",
        "",
    ]

    if result.warnings:
        lines += [f"## {t('common.warnings', lang)}", ""]
        for w in result.warnings:
            lines.append(f"- {w}")
        lines.append("")

    lines += ["---", f"*AUTOMATION_FACTORY script_protocol_generator.py — {ts}*"]

    out_dir = project_path / "_output"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{result.test_type}_TEST_PROTOCOL_{ts_file}.md"
    out_path.write_text("\n".join(lines), encoding="utf-8")
    result.md_path = out_path
    return out_path


def write_protocol_xlsx(result: ProtocolResult, project_path: Path) -> Optional[Path]:
    """Write the protocol to an Excel file (openpyxl required)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None

    lang = result.lang
    ts_file = datetime.now().strftime("%Y%m%d_%H%M")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{result.test_type} Protocol"

    # Title
    ws.merge_cells("A1:H1")
    title = t("common.sat_title" if result.test_type == "SAT" else "common.fat_title", lang)
    ws["A1"] = f"{title} — {result.project_name}"
    ws["A1"].font = Font(bold=True, size=14)

    # Column headers
    headers = [t("col.test_id", lang), t("col.description", lang)[:14],
               t("col.description", lang), t("col.tag", lang),
               t("col.expected", lang), t("col.ref", lang),
               t("col.actual", lang), t("col.signature", lang)]
    headers[1] = "Section"
    col_widths = [12, 14, 50, 20, 35, 14, 20, 16]
    hdr_fill = PatternFill("solid", fgColor="1F497D")
    for c, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hdr_fill
        cell.alignment = Alignment(wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = w

    section_colors = {
        "PreCheck": "D9E1F2",
        "IO":       "E2EFDA",
        "Function": "FFF2CC",
        "Alarm":    "FCE4D6",
    }
    for row_idx, item in enumerate(result.items, start=3):
        fill_c = section_colors.get(item.section, "FFFFFF")
        fill = PatternFill("solid", fgColor=fill_c)
        vals = [item.id, item.section, item.description, item.tag,
                item.expected, item.ref, item.result_placeholder, ""]
        for c, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=c, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True)

    out_dir = project_path / "_output"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{result.test_type}_TEST_PROTOCOL_{ts_file}.xlsx"
    wb.save(str(out_path))
    result.xlsx_path = out_path
    return out_path


def run_protocol(
    project_path: Path,
    test_type: str = "FAT",
    write_excel: bool = False,
    lang: str = DEFAULT_LANG,
    pdf: bool = False,
) -> ProtocolResult:
    """Full pipeline wrapper for a single protocol type."""
    result = generate_protocol(project_path, test_type=test_type, lang=lang)
    write_protocol_md(result, project_path)
    if write_excel:
        write_protocol_xlsx(result, project_path)
    if pdf and result.md_path:
        try:
            from pdf_common import markdown_to_pdf
            md_text = result.md_path.read_text(encoding="utf-8")
            title = t("common.sat_title" if result.test_type == "SAT" else "common.fat_title", result.lang)
            result.pdf_path = markdown_to_pdf(
                md_text, result.md_path.with_suffix(".pdf"),
                f"{title} — {result.project_name}",
            )
        except Exception as exc:
            # Fail-safe: MD remains; the PDF failure is loud, never silent.
            result.warnings.append(f"{t('status.pdf_failed', result.lang)} ({exc})")
    return result


def run_protocol_set(
    project_path: Path,
    test_type: str = "FAT",
    write_excel: bool = False,
    lang: str = DEFAULT_LANG,
    pdf: bool = False,
) -> list[ProtocolResult]:
    """FAT, SAT or BOTH — BOTH produces two separate documents."""
    tt = (test_type or "FAT").upper()
    if tt == "BOTH":
        types = ("FAT", "SAT")
    elif tt in ("FAT", "SAT"):
        types = (tt,)
    else:
        raise ValueError(f"Unknown protocol type {test_type!r} — use FAT, SAT or BOTH.")
    return [
        run_protocol(project_path, test_type=x, write_excel=write_excel,
                     lang=lang, pdf=pdf)
        for x in types
    ]


def format_protocol_summary(result: ProtocolResult) -> str:
    lines = [f"{result.test_type} Test Protocol Summary", ""]
    lines.append(f"  Project     : {result.project_name}")
    lines.append(f"  Language    : {result.lang}")
    lines.append(f"  Total tests : {len(result.items)}")
    lines.append(f"    Pre-Check   : {len(result.precheck_items)}")
    lines.append(f"    IO Test     : {len(result.io_items)}")
    lines.append(f"    Function    : {len(result.function_items)}")
    lines.append(f"    Alarm Test  : {len(result.alarm_items)}")
    # Log hygiene: file NAME only — the full path can carry the customer
    # project folder and must not reach stdout/logs.
    if result.md_path:
        lines.append(f"  MD Output   : {result.md_path.name}")
    if result.xlsx_path:
        lines.append(f"  Excel Output: {result.xlsx_path.name}")
    if result.pdf_path:
        lines.append(f"  PDF Output  : {result.pdf_path.name}")
    if result.warnings:
        lines.append("")
        for w in result.warnings:
            lines.append(f"  {w}")
    return "\n".join(lines)


# -- CLI ----------------------------------------------------------------------

def main():
    force_utf8_stdout()
    import argparse
    p = argparse.ArgumentParser(description="FAT/SAT Test Protocol Generator")
    p.add_argument("--project", metavar="PROJECT_PATH", required=True)
    p.add_argument("--type", choices=["FAT", "SAT", "BOTH"], default="FAT")
    p.add_argument("--lang", choices=["de", "en", "tr"], default=DEFAULT_LANG)
    p.add_argument("--pdf", action="store_true", help="Also produce PDF")
    p.add_argument("--excel", action="store_true", help="Also produce Excel")
    args = p.parse_args()

    results = run_protocol_set(
        Path(args.project), test_type=args.type, write_excel=args.excel,
        lang=args.lang, pdf=args.pdf,
    )
    for result in results:
        print(format_protocol_summary(result))


if __name__ == "__main__":
    main()

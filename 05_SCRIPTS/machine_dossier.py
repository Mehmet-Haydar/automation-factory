#!/usr/bin/env python3
"""machine_dossier.py — Machine Dossier generator (deterministic).

Six engineer/customer-facing pages per machine, generated from the proven
legacy extraction. They live on the APPROVAL side (metadata/machine_dossier/)
because their purpose is: analyze → engineer edits the decision columns →
gate sign-off → code generation. After approval the same pages go into the
handover package as PDF/XLSX so customer and engineer mark up the SAME visual.

  01_operator_flow.svg    "what happens when START is pressed" (PAP style)
  02_block_structure.svg  OB→PB/FB call structure; roles inferred from writes
  03_grafcet_<n>.svg      one DIN EN 60848-style page per proven step chain
  04_decision_table.xlsx  device rows; the DECISION columns belong to the
     (+ .md mirror)       engineer and are NEVER auto-filled
  05_plant_summary.md     deterministic counts only
  06_ce_matrix.xlsx/.md   cause→effect marks from the proven rail equations
  state_table.json        the single data source every view renders from

Look & feel comes ONLY from theme.json at the factory root (single file —
swap the palette without touching code). AI is never called here; signals
without a symbol name stay "❓" instead of being guessed.
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

from legacy_enrich import load_symbols  # type: ignore
from s5_logic_extract import (  # type: ignore
    And, NetworkLogic, Not, Or, Var, extract_project_logic, render_expr,
)
from sequence_map import build_chains  # type: ignore

_FACTORY_ROOT = Path(__file__).resolve().parents[1]

_DEFAULT_THEME = {
    "page_bg": "#FFFFFF", "ink": "#1B2430", "ink_soft": "#5A6672",
    "step_fill": "#F2F7F9", "step_stroke": "#0F6E8C", "step_stroke_width": 2.0,
    "initial_double_gap": 5, "action_fill": "#EAF4F7",
    "action_stroke": "#0F6E8C", "action_stroke_width": 1.0,
    "wire": "#3C4854", "wire_width": 1.4,
    "transition_bar": "#1B2430", "transition_bar_width": 3.5,
    "cond_color": "#1B2430", "cond_addr_color": "#0F6E8C",
    "global_band_fill": "#FBEDEC", "global_band_stroke": "#B3423E",
    "global_band_text": "#8E3532",
    "font_family": "Segoe UI, Arial, sans-serif",
    "mono_family": "Consolas, Cascadia Code, monospace",
    "fs_title": 19, "fs_subtitle": 12, "fs_step_num": 15,
    "fs_action_title": 12.5, "fs_action_sub": 10.5, "fs_cond": 11,
    "fs_footer": 10, "corner_radius": 3,
}


def load_theme() -> dict:
    """theme.json at the factory root; missing keys fall back to defaults."""
    theme = dict(_DEFAULT_THEME)
    fp = _FACTORY_ROOT / "theme.json"
    try:
        theme.update(json.loads(fp.read_text(encoding="utf-8")))
    except Exception:
        pass
    return theme


@dataclass
class DossierSummary:
    chains: int = 0
    steps: int = 0
    blocks: int = 0
    signals: int = 0
    files: list = field(default_factory=list)
    warnings: list = field(default_factory=list)
    out_dir: Path | None = None


# ---------------------------------------------------------------------------
# expression helpers
# ---------------------------------------------------------------------------

def _atoms(e, neg: bool = False) -> list[tuple[bool, str]]:
    """In-order (negated, operand) leaves of an expression tree."""
    if e is None:
        return []
    if isinstance(e, Var):
        return [(neg, e.operand)]
    if isinstance(e, Not):
        return _atoms(e.a, not neg)
    if isinstance(e, (And, Or)):
        return _atoms(e.a, neg) + _atoms(e.b, neg)
    return []


def _dedup(pairs: list[tuple[bool, str]]) -> list[tuple[bool, str]]:
    seen, out = set(), []
    for p in pairs:
        if p not in seen:
            seen.add(p)
            out.append(p)
    return out


def _and_factors(e) -> list:
    """Top-level AND factors of an expression (AND is the transition glue)."""
    if isinstance(e, And):
        return _and_factors(e.a) + _and_factors(e.b)
    return [] if e is None else [e]


def _or_terms(e) -> list:
    if isinstance(e, Or):
        return _or_terms(e.a) + _or_terms(e.b)
    return [e]


def _as_atom(e) -> tuple[bool, str] | None:
    """(negated, operand) if the subtree is a bare/negated variable."""
    neg = False
    while isinstance(e, Not):
        neg = not neg
        e = e.a
    return (neg, e.operand) if isinstance(e, Var) else None


def _cond_groups(e, hide: set[str], names: dict[str, str]) -> list[dict]:
    """AST-aware transition condition — keeps the OR structure that a flat
    atom list destroys. Each top-level AND factor becomes one group:

      atom  {kind:"atom", neg, addr, label}   hidden if addr ∈ hide
      or    {kind:"or", atoms:[…]}            pure OR of atoms — alternative
            entries into the step (bypass/jog/retrigger). Atoms inside an OR
            are NEVER hidden: dropping one alternative changes the meaning.
      expr  {kind:"expr", text}               anything deeper — shown as the
            exact rendered expression instead of a silently wrong flat list.
    """
    groups: list[dict] = []
    for f in _and_factors(e):
        atom = _as_atom(f)
        if atom is not None:
            neg, op = atom
            if op in hide:
                continue
            groups.append({"kind": "atom", "neg": neg, "addr": op,
                           "label": names.get(op, "")})
            continue
        if isinstance(f, Or):
            terms = [_as_atom(t) for t in _or_terms(f)]
            if all(t is not None for t in terms):
                atoms = [{"neg": n, "addr": o, "label": names.get(o, "")}
                         for n, o in _dedup([t for t in terms if t])]
                groups.append({"kind": "or", "atoms": atoms})
                continue
        groups.append({"kind": "expr", "text": render_expr(f, names)})
    return groups


# ---------------------------------------------------------------------------
# state table — the single data source
# ---------------------------------------------------------------------------

def _threads(latches: dict) -> list[list[str]]:
    """Split the latch graph into linear THREADS using the classic S5
    Schrittkette signature: a → b is a thread edge only when a appears
    positive in b's SET **and** b appears positive in a's RESET (step k+1
    resets step k). Cross-chain triggers (a in b's SET without the reset
    link) do NOT merge threads — that is exactly the 1N/2N/3N split which
    plain connected components get wrong (live finding, 2026-07-06)."""
    def _pos(e) -> set[str]:
        return {op for neg, op in _dedup(_atoms(e)) if not neg}

    fwd: dict[str, set[str]] = {m: set() for m in latches}
    for b, (_nl, coil_b) in latches.items():
        for a in _pos(coil_b.set_cond):
            if a in latches and a != b and b in _pos(latches[a][1].reset_cond):
                fwd[a].add(b)

    undirected: dict[str, set[str]] = {m: set() for m in latches}
    for a, targets in fwd.items():
        for b in targets:
            undirected[a].add(b)
            undirected[b].add(a)

    seen: set[str] = set()
    threads: list[list[str]] = []
    for start in sorted(latches):
        if start in seen:
            continue
        comp, stack = [], [start]
        while stack:
            n = stack.pop()
            if n in seen:
                continue
            seen.add(n)
            comp.append(n)
            stack.extend(sorted(undirected[n] - seen))
        if len(comp) < 3:
            continue
        comp_set = set(comp)
        preds = {m: {a for a in comp_set if m in fwd[a]} for m in comp}
        ordered: list[str] = []
        left = set(comp)
        while left:
            ready = sorted(m for m in left if preds[m] <= set(ordered))
            if not ready:               # cycle — take lowest, stay honest
                ready = [sorted(left)[0]]
            ordered.append(ready[0])
            left.discard(ready[0])
        threads.append(ordered)
    return threads


def build_state_table(project_root: Path) -> dict:
    root = Path(project_root)
    nets = extract_project_logic(root)
    latches, _components = build_chains(nets)
    chains = _threads(latches)
    names = load_symbols(root / "_raw" / "legacy_code")

    chain_dicts = []
    for ci, chain in enumerate(chains, 1):
        members = set(chain)
        # rails: non-member M vars present in (almost) every SET / RESET
        set_counts: dict[str, int] = {}
        reset_counts: dict[str, int] = {}
        for step in chain:
            _nl, coil = latches[step]
            for neg, op in _dedup(_atoms(coil.set_cond)):
                if not neg and op.startswith("M") and op not in members:
                    set_counts[op] = set_counts.get(op, 0) + 1
            for neg, op in _dedup(_atoms(coil.reset_cond)):
                if not neg and op.startswith("M") and op not in members:
                    reset_counts[op] = reset_counts.get(op, 0) + 1
        thresh = max(2, int(len(chain) * 0.8))
        enable_rails = sorted(op for op, n in set_counts.items() if n >= thresh)
        reset_rails = sorted(op for op, n in reset_counts.items() if n >= thresh)

        steps = []
        for step in chain:
            nl, coil = latches[step]
            hide = members | set(enable_rails) | set(reset_rails)
            conds = [
                {"neg": neg, "addr": op, "label": names.get(op, "")}
                for neg, op in _dedup(_atoms(coil.set_cond))
                if op not in hide
            ]
            steps.append({
                "merker": step,
                "name": names.get(step, ""),
                "source": f"{nl.block}/N{nl.network}",
                "proven_vectors": nl.verified_vectors,
                "set": render_expr(coil.set_cond, names),
                "reset": render_expr(coil.reset_cond, names),
                "conditions": conds,
                "cond_groups": _cond_groups(coil.set_cond, hide, names),
            })
        chain_dicts.append({
            "id": f"chain_{ci}",
            "label": steps[0]["name"] or chain[0],
            "members": chain,
            "enable_rails": [
                {"addr": r, "label": names.get(r, "")} for r in enable_rails],
            "reset_rails": [
                {"addr": r, "label": names.get(r, "")} for r in reset_rails],
            "steps": steps,
        })

    # cross-chain triggers: step of chain A positive in SET of chain B's step
    triggers = []
    owner = {m: ci for ci, ch in enumerate(chains, 1) for m in ch}
    for ci, chain in enumerate(chains, 1):
        for step in chain:
            _nl, coil = latches[step]
            for neg, op in _dedup(_atoms(coil.set_cond)):
                if neg or op not in owner or owner[op] == ci:
                    continue
                triggers.append({
                    "from_chain": f"chain_{owner[op]}", "from_step": op,
                    "from_label": names.get(op, ""),
                    "to_chain": f"chain_{ci}", "to_step": step,
                })

    return {
        "schema": "state_table/v1",
        "generated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "networks": len(nets),
        "networks_proven": sum(1 for n in nets if n.parsed),
        "chains": chain_dicts,
        "triggers": triggers,
        "provenance": ("deterministic S/R extraction, self-proven on random "
                       "vectors per network; DRAFT_UNVERIFIED — engineer "
                       "confirms intent"),
    }


# ---------------------------------------------------------------------------
# OB → block call structure
# ---------------------------------------------------------------------------

_CALL_RE = re.compile(
    r"^\s*(?:JU|JC|SPA|SPB|CALL|BA|BAB)\s+(?:\"?)(PB|FB|OB|FC|SB)\s?(\d{1,3})",
    re.I,
)


def parse_block_calls(legacy_dir: Path) -> list[tuple[str, str]]:
    """[(caller, callee), …] in source order, from OB*.AWL files."""
    calls: list[tuple[str, str]] = []
    if not legacy_dir.is_dir():
        return calls
    for fp in sorted(legacy_dir.iterdir()):
        if not fp.is_file() or fp.suffix.lower() != ".awl":
            continue
        stem = fp.stem.upper()
        if not stem.startswith("OB"):
            continue
        try:
            text = fp.read_text(encoding="utf-8", errors="replace")
        except Exception:
            continue
        for line in text.splitlines():
            m = _CALL_RE.match(line)
            if m:
                calls.append((stem, f"{m.group(1).upper()}{m.group(2)}"))
    return calls


def _block_roles(nets: list[NetworkLogic], state: dict,
                 names: dict[str, str]) -> dict[str, str]:
    """block -> short role text, inferred ONLY from what the block writes."""
    writes: dict[str, set[str]] = {}
    for nl in nets:
        if nl.parsed:
            writes.setdefault(nl.block.upper(), set()).update(nl.coils.keys())

    chain_of = {m: ch["label"] for ch in state["chains"] for m in ch["members"]}
    rails = {r["addr"] for ch in state["chains"]
             for r in ch["enable_rails"] + ch["reset_rails"]}

    roles: dict[str, str] = {}
    for block, ops in writes.items():
        chain_hits = sorted({chain_of[o] for o in ops if o in chain_of})
        if chain_hits:
            roles[block] = f"step logic: {', '.join(chain_hits)[:44]}"
        elif ops & rails:
            roles[block] = "enable rails (" + ", ".join(sorted(ops & rails)) + ")"
        elif any(o.startswith("Q") for o in ops):
            roles[block] = "output driving (Q coils)"
        elif ops:
            first = sorted(ops)[0]
            label = names.get(first, "")
            roles[block] = f"writes {first}" + (f" — {label[:30]}" if label else "")
    return roles


# ---------------------------------------------------------------------------
# SVG primitives (theme-driven)
# ---------------------------------------------------------------------------

def _esc(s: str) -> str:
    return (s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            .replace('"', "&quot;"))


def _short(s: str, n: int) -> str:
    return s if len(s) <= n else s[: n - 1] + "…"


class _Svg:
    def __init__(self, theme: dict, width: int):
        self.t = theme
        self.w = width
        self.body: list[str] = []

    def text(self, x, y, s, fs, fill, family=None, anchor="start",
             weight="", transform=""):
        fam = family or self.t["font_family"]
        w = f' font-weight="{weight}"' if weight else ""
        tr = f' transform="{transform}"' if transform else ""
        self.body.append(
            f'<text x="{x}" y="{y}" font-size="{fs}" fill="{fill}" '
            f'font-family="{_esc(fam)}" text-anchor="{anchor}"{w}{tr}>'
            f'{_esc(s)}</text>')

    def rect(self, x, y, w, h, fill, stroke, sw, rx=None, dash=""):
        rx = self.t["corner_radius"] if rx is None else rx
        d = f' stroke-dasharray="{dash}"' if dash else ""
        self.body.append(
            f'<rect x="{x}" y="{y}" width="{w}" height="{h}" rx="{rx}" '
            f'fill="{fill}" stroke="{stroke}" stroke-width="{sw}"{d}/>')

    def line(self, x1, y1, x2, y2, stroke, sw, dash="", arrow=False):
        d = f' stroke-dasharray="{dash}"' if dash else ""
        a = ' marker-end="url(#arr)"' if arrow else ""
        self.body.append(
            f'<line x1="{x1}" y1="{y1}" x2="{x2}" y2="{y2}" '
            f'stroke="{stroke}" stroke-width="{sw}"{d}{a}/>')

    def poly(self, points: str, stroke, sw, fill="none", dash="", arrow=False):
        d = f' stroke-dasharray="{dash}"' if dash else ""
        a = ' marker-end="url(#arr)"' if arrow else ""
        self.body.append(
            f'<polyline points="{points}" fill="{fill}" stroke="{stroke}" '
            f'stroke-width="{sw}"{d}{a}/>')

    def diamond(self, cx, cy, hw, hh, fill, stroke, sw):
        self.body.append(
            f'<polygon points="{cx},{cy-hh} {cx+hw},{cy} {cx},{cy+hh} '
            f'{cx-hw},{cy}" fill="{fill}" stroke="{stroke}" '
            f'stroke-width="{sw}"/>')

    def render(self, height: int) -> str:
        t = self.t
        return (
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{self.w}" '
            f'height="{height}" viewBox="0 0 {self.w} {height}" '
            f'font-family="{_esc(t["font_family"])}">'
            f'<rect width="{self.w}" height="{height}" fill="{t["page_bg"]}"/>'
            '<defs><marker id="arr" markerWidth="9" markerHeight="9" refX="8" '
            f'refY="4.5" orient="auto"><path d="M0,0 L9,4.5 L0,9 z" '
            f'fill="{t["wire"]}"/></marker></defs>'
            + "".join(self.body) + "</svg>")


def _draft_band(svg: _Svg, y: int, lines: list[str]) -> int:
    t = svg.t
    h = 16 + len(lines) * 17
    svg.rect(28, y, svg.w - 56, h, t["global_band_fill"],
             t["global_band_stroke"], 1.2, dash="6 4")
    yy = y + 19
    for ln in lines:
        svg.text(40, yy, ln, t["fs_cond"], t["global_band_text"],
                 family=t["mono_family"])
        yy += 17
    return y + h


# ---------------------------------------------------------------------------
# view 03 — GRAFCET per chain
# ---------------------------------------------------------------------------

def render_grafcet(chain: dict, machine: str, theme: dict) -> str:
    W = 980
    svg = _Svg(theme, W)
    t = theme
    x_line, step_w, x_action, action_w = 110, 46, 200, 430

    svg.text(28, 44, f"GRAFCET — {_short(chain['label'], 60)}",
             t["fs_title"], t["ink"], weight="600")
    svg.text(28, 64, f"{machine} · proven S/R extraction · DRAFT — engineer "
             "approval required · DIN EN 60848 style",
             t["fs_subtitle"], t["ink_soft"])

    rails = " · ".join(
        f"{r['addr']} {r['label']}".strip()
        for r in chain["enable_rails"]) or "—"
    resets = " · ".join(
        f"{r['addr']} {r['label']}".strip()
        for r in chain["reset_rails"]) or "—"
    y = _draft_band(svg, 78, [
        f"GLOBAL: every transition ∧ {_short(rails, 96)}",
        f"RESET rail: {_short(resets, 100)}",
    ]) + 16

    # initial step (double square)
    sx = x_line - step_w / 2
    svg.rect(sx, y, step_w, step_w, t["step_fill"], t["step_stroke"],
             t["step_stroke_width"], rx=2)
    g = t["initial_double_gap"]
    svg.rect(sx + g, y + g, step_w - 2 * g, step_w - 2 * g, "none",
             t["step_stroke"], t["step_stroke_width"] * 0.8, rx=1)
    svg.text(x_line, y + step_w / 2 + 5, "0", t["fs_step_num"], t["ink"],
             anchor="middle", weight="700", family=t["mono_family"])
    svg.line(sx + step_w, y + step_w / 2, x_action, y + step_w / 2,
             t["wire"], t["wire_width"])
    svg.rect(x_action, y + step_w / 2 - 16, action_w, 32, t["action_fill"],
             t["action_stroke"], t["action_stroke_width"])
    svg.text(x_action + 10, y + step_w / 2 + 4,
             "initial — chain idle, waiting for enable",
             t["fs_action_title"], t["ink"])
    loop_top = y + step_w / 2
    y += step_w

    def _atom_line(x, yy, a, prefix=""):
        pre = "¬" if a["neg"] else ""
        px = (f'<tspan fill="{t["ink_soft"]}">{_esc(prefix)}</tspan>'
              if prefix else "")
        svg.body.append(
            f'<text x="{x}" y="{yy}" font-size="{t["fs_cond"]}" '
            f'font-family="{_esc(t["mono_family"])}">{px}'
            f'<tspan fill="{t["cond_addr_color"]}">'
            f'{_esc(pre + a["addr"])}</tspan>'
            f'<tspan fill="{t["cond_color"]}"> '
            f'{_esc(_short(a["label"] or "❓", 42))}</tspan></text>')

    for i, st in enumerate(chain["steps"], 1):
        # AST-aware groups (2026-07-06 fix: OR alternatives no longer flatten
        # into the AND list); old state_table.json falls back to flat atoms.
        groups = st.get("cond_groups") or [
            {"kind": "atom", **c} for c in st["conditions"]]
        n = max(1, sum(len(g["atoms"]) if g["kind"] == "or" else 1
                       for g in groups))
        zone = 30 + n * 15 + 10
        svg.line(x_line, y, x_line, y + zone, t["wire"], t["wire_width"])
        bar_y = y + zone / 2
        svg.line(x_line - 14, bar_y, x_line + 14, bar_y,
                 t["transition_bar"], t["transition_bar_width"])
        cy = bar_y - (n - 1) * 7.5 + 4
        if not groups:
            svg.text(x_line + 26, cy, "(previous step + global rails only)",
                     t["fs_cond"], t["ink_soft"])
        for g in groups:
            if g["kind"] == "or":
                # alternatives bound by a bracket; "∨" on lines 2..n makes
                # the any-one-of reading explicit (DIN EN 60848 boolean text)
                top = cy
                for j, a in enumerate(g["atoms"]):
                    _atom_line(x_line + 40, cy, a, prefix="∨ " if j else "")
                    cy += 15
                svg.line(x_line + 32, top - 9, x_line + 32, cy - 21,
                         t["cond_addr_color"], 1.2)
            elif g["kind"] == "expr":
                svg.text(x_line + 26, cy, _short(g["text"], 64), t["fs_cond"],
                         t["cond_color"], family=t["mono_family"])
                cy += 15
            else:
                _atom_line(x_line + 26, cy, g)
                cy += 15
        y += zone

        svg.rect(sx, y, step_w, step_w, t["step_fill"], t["step_stroke"],
                 t["step_stroke_width"], rx=2)
        svg.text(x_line, y + step_w / 2 + 5, str(i), t["fs_step_num"],
                 t["ink"], anchor="middle", weight="700",
                 family=t["mono_family"])
        svg.line(sx + step_w, y + step_w / 2, x_action, y + step_w / 2,
                 t["wire"], t["wire_width"])
        svg.rect(x_action, y + step_w / 2 - 19, action_w, 38,
                 t["action_fill"], t["action_stroke"],
                 t["action_stroke_width"])
        svg.text(x_action + 10, y + step_w / 2 - 3,
                 _short(st["name"] or st["merker"], 56),
                 t["fs_action_title"], t["ink"], weight="600")
        svg.text(x_action + 10, y + step_w / 2 + 13,
                 f"{st['merker']} · {st['source']} · proven/"
                 f"{st['proven_vectors']}",
                 t["fs_action_sub"], t["ink_soft"], family=t["mono_family"])
        y += step_w

    # loop back
    svg.line(x_line, y, x_line, y + 24, t["wire"], t["wire_width"])
    bar_y = y + 24
    svg.line(x_line - 14, bar_y, x_line + 14, bar_y, t["transition_bar"],
             t["transition_bar_width"])
    svg.text(x_line + 26, bar_y + 4, "cycle complete — back to initial",
             t["fs_cond"], t["ink"], family=t["mono_family"])
    lx = 46
    svg.poly(f"{x_line},{bar_y+12} {x_line},{bar_y+28} {lx},{bar_y+28} "
             f"{lx},{loop_top} {sx},{loop_top}", t["wire"], t["wire_width"],
             arrow=True)

    H = int(bar_y + 64)
    svg.text(28, H - 14, "source: state_table.json · look: theme.json · "
             "❓ = no symbol name in the legacy export (never guessed)",
             t["fs_footer"], t["ink_soft"])
    return svg.render(H)


# ---------------------------------------------------------------------------
# view 01 — operator flow ("what happens when START is pressed")
# ---------------------------------------------------------------------------

def render_operator_flow(state: dict, machine: str, theme: dict) -> str:
    W = 980
    svg = _Svg(theme, W)
    t = theme
    cx = 470

    svg.text(28, 40, "OPERATOR FLOW — what happens on START?",
             t["fs_title"], t["ink"], weight="600")
    svg.text(28, 60, f"{machine} · deterministic skeleton from the proven "
             "chain-start equation · DRAFT — engineer approval required",
             t["fs_subtitle"], t["ink_soft"])

    chains = state["chains"]
    y = 84
    if not chains:
        svg.text(28, 110, "No proven step chain found — nothing to draw "
                 "(honest refusal).", 13, t["ink"])
        return svg.render(150)

    first = chains[0]
    start_conds = first["steps"][0]["conditions"][:6]
    folded = len(first["steps"][0]["conditions"]) - len(start_conds)

    # start pill
    svg.rect(cx - 160, y, 320, 42, t["step_fill"], t["step_stroke"], 2, rx=21)
    svg.text(cx, y + 26, "▶ START / mode selected", 13, t["ink"],
             anchor="middle", weight="600")
    y += 42

    for c in start_conds:
        svg.line(cx, y, cx, y + 22, t["wire"], t["wire_width"], arrow=True)
        y += 22
        hh = 34
        svg.diamond(cx, y + hh, 165, hh, t["action_fill"], t["step_stroke"], 1.4)
        label = c["label"] or "❓"
        q = ("NOT " if c["neg"] else "") + _short(label, 34) + " ?"
        svg.text(cx, y + hh - 2, q, 11.5, t["ink"], anchor="middle",
                 weight="600")
        svg.text(cx, y + hh + 14, c["addr"], 10, t["ink_soft"],
                 anchor="middle", family=t["mono_family"])
        # NEIN branch
        svg.line(cx + 165, y + hh, cx + 255, y + hh, t["wire"],
                 t["wire_width"], arrow=True)
        svg.text(cx + 175, y + hh - 6, "NO", 10.5, t["global_band_text"])
        svg.rect(cx + 258, y + hh - 22, 220, 44, t["global_band_fill"],
                 t["global_band_stroke"], 1.1)
        svg.text(cx + 268, y + hh - 4, f"check: {_short(label or '❓', 28)}",
                 10.5, t["global_band_text"])
        svg.text(cx + 268, y + hh + 12, c["addr"], 10,
                 t["global_band_text"], family=t["mono_family"])
        svg.text(cx + 8, y + 2 * hh + 12, "YES", 10.5, "#2E7D46")
        y += 2 * hh

    if folded > 0:
        svg.line(cx, y, cx, y + 22, t["wire"], t["wire_width"], arrow=True)
        y += 22
        svg.rect(cx - 165, y, 330, 34, t["page_bg"], t["ink_soft"], 1.1,
                 dash="4 3")
        svg.text(cx, y + 21, f"+ {folded} more conditions — see GRAFCET page",
                 11, t["ink_soft"], anchor="middle")
        y += 34

    # enable + chain start
    svg.line(cx, y, cx, y + 24, t["wire"], t["wire_width"], arrow=True)
    y += 24
    rails = " · ".join(f"{r['addr']} {r['label']}".strip()
                       for r in first["enable_rails"]) or "chain enable"
    svg.rect(cx - 190, y, 380, 48, t["step_fill"], t["step_stroke"], 2.4)
    svg.text(cx, y + 20, "CHAIN ENABLE established", 12.5, t["cond_addr_color"],
             anchor="middle", weight="700")
    svg.text(cx, y + 38, _short(rails, 58), 10.5, t["ink_soft"],
             anchor="middle", family=t["mono_family"])
    y += 48
    svg.line(cx, y, cx, y + 24, t["wire"], t["wire_width"], arrow=True)
    y += 24
    svg.rect(cx - 190, y, 380, 48, t["action_fill"], t["step_stroke"], 1.6)
    svg.text(cx, y + 20, f"chain starts: {_short(first['label'], 40)}",
             12, t["ink"], anchor="middle", weight="600")
    svg.text(cx, y + 38, f"{len(first['steps'])} steps · GRAFCET page 1",
             10.5, t["ink_soft"], anchor="middle", family=t["mono_family"])
    y += 48

    # cross-chain triggers
    trig = [tr for tr in state["triggers"] if tr["from_chain"] == first["id"]]
    if trig:
        by_step: dict[str, list] = {}
        for tr in trig:
            by_step.setdefault(tr["from_step"], []).append(tr)
        step_id, tos = sorted(by_step.items(),
                              key=lambda kv: -len(kv[1]))[0]
        label = next((tr["from_label"] for tr in tos if tr["from_label"]),
                     step_id)
        targets = sorted({tr["to_chain"] for tr in tos})
        tgt_labels = [
            _short(next(c["label"] for c in chains if c["id"] == tc), 30)
            for tc in targets]
        svg.line(cx, y, cx, y + 24, t["wire"], t["wire_width"], arrow=True)
        y += 24
        svg.rect(cx - 190, y, 380, 52, t["action_fill"], t["step_stroke"], 1.4)
        svg.text(cx, y + 20, f"trigger: {_short(label, 44)}", 11.5, t["ink"],
                 anchor="middle", weight="600")
        svg.text(cx, y + 38, "starts → " + _short(" + ".join(tgt_labels), 52),
                 10.5, t["ink_soft"], anchor="middle")
        y += 52

    svg.line(cx, y, cx, y + 24, t["wire"], t["wire_width"], arrow=True)
    y += 24
    svg.rect(cx - 160, y, 320, 40, t["step_fill"], t["step_stroke"], 2, rx=20)
    svg.text(cx, y + 25, "cycle complete — repeats", 12.5, t["ink"],
             anchor="middle", weight="600")
    y += 40

    y = _draft_band(svg, y + 18, [
        "Interlocks (e-stop, motor protection, …) are NOT on this page by "
        "design → 06_ce_matrix",
        "step details → 03_grafcet pages · devices → 04_decision_table",
    ])
    H = y + 40
    svg.text(28, H - 14, "deterministic skeleton — the friendly wording is "
             "the engineer's to edit (DRAFT)", t["fs_footer"], t["ink_soft"])
    return svg.render(H)


# ---------------------------------------------------------------------------
# view 02 — block structure
# ---------------------------------------------------------------------------

def render_block_structure(calls: list[tuple[str, str]], roles: dict,
                           machine: str, theme: dict) -> str:
    W = 980
    svg = _Svg(theme, W)
    t = theme

    svg.text(28, 40, "BLOCK STRUCTURE — OB → PB/FB (rough map)",
             t["fs_title"], t["ink"], weight="600")
    svg.text(28, 60, f"{machine} · source: call statements in the OB sources "
             "(order preserved) · ❓ = role not derivable from code",
             t["fs_subtitle"], t["ink_soft"])

    callers = []
    for c, _ in calls:
        if c not in callers:
            callers.append(c)
    if not callers:
        svg.text(28, 110, "No OB call statements found — nothing to draw "
                 "(honest refusal).", 13, t["ink"])
        return svg.render(150)

    y = 96
    for caller in callers:
        callees = [b for a, b in calls if a == caller]
        box_h = 24
        col_h = len(callees) * (box_h + 8) - 8
        oy = y + max(0, col_h / 2 - 30)
        svg.rect(60, oy, 150, 60, t["step_fill"], t["step_stroke"], 2.4)
        svg.text(135, oy + 26, caller, 14, t["cond_addr_color"],
                 anchor="middle", weight="700")
        svg.text(135, oy + 44, "cyclic main" if caller == "OB1" else "system",
                 10.5, t["ink_soft"], anchor="middle")
        by = y
        for callee in callees:
            role = roles.get(callee, "")
            unknown = not role
            svg.line(210, oy + 30, 292, by + box_h / 2, t["wire"], 1.1,
                     arrow=True)
            svg.rect(295, by, 620, box_h, t["action_fill"] if not unknown
                     else t["page_bg"],
                     t["step_stroke"] if not unknown
                     else t["global_band_stroke"],
                     1.2, dash="" if not unknown else "4 3")
            label = f"{callee} — {role}" if role else f"{callee} — ❓"
            svg.text(305, by + 16.5, _short(label, 88), 11,
                     t["ink"] if not unknown else t["global_band_text"],
                     family=t["mono_family"])
            by += box_h + 8
        y = max(by, oy + 60) + 18

    y = _draft_band(svg, y, [
        "role inference: what each block WRITES in the proven extraction — "
        "no AI, no guessing",
    ])
    H = y + 30
    return svg.render(H)


# ---------------------------------------------------------------------------
# tables (xlsx + md)
# ---------------------------------------------------------------------------

_DECISION_HEADERS = [
    "Address", "Name", "Type", "Equipment", "Function (from code/symbols)",
    "Current solution", "DECISION: new solution (ENGINEER)",
    "Code impact (ENGINEER)",
]


def _signals(root: Path) -> list[dict]:
    try:
        from iec_tag_generator import parse_rd01_signals  # type: ignore
        return parse_rd01_signals(Path(root))
    except Exception:
        return []


def _write_xlsx(path: Path, headers: list[str], rows: list[list[str]],
                title: str, note: str) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except Exception:
        return False
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append([note])
    ws["A1"].font = Font(italic=True, size=9)
    ws.append(headers)
    fill = PatternFill("solid", fgColor="DDE8EC")
    for c in ws[2]:
        c.font = Font(bold=True)
        c.fill = fill
    for r in rows:
        ws.append(r)
    for i, h in enumerate(headers, 1):
        width = max(12, min(46, max([len(h)] + [len(str(r[i - 1]))
                    for r in rows] or [12]) + 2))
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = width
    ws.freeze_panes = "A3"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return True


def _write_md_table(path: Path, title: str, note: str, headers: list[str],
                    rows: list[list[str]]) -> None:
    L = [f"# {title}", "", f"> {note}", ""]
    L.append("| " + " | ".join(headers) + " |")
    L.append("|" + "---|" * len(headers))
    for r in rows:
        L.append("| " + " | ".join(str(c).replace("|", "\\|") for c in r) + " |")
    path.write_text("\n".join(L) + "\n", encoding="utf-8")


def load_decisions(root: Path) -> dict[str, dict]:
    """Engineer-owned modernization decisions, keyed by signal address.

    They live in their own file (decisions.json) precisely so that
    re-generating the dossier can NEVER erase what the engineer wrote —
    the generated table is a merge of deterministic rows + this file."""
    fp = Path(root) / "metadata" / "machine_dossier" / "decisions.json"
    try:
        data = json.loads(fp.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def save_decisions(root: Path, entries: dict[str, dict]) -> int:
    """Merge engineer entries into decisions.json and refresh the two
    generated table files. Empty decision+impact removes the entry."""
    root = Path(root)
    current = load_decisions(root)
    for addr, val in (entries or {}).items():
        decision = str((val or {}).get("decision", "")).strip()
        impact = str((val or {}).get("impact", "")).strip()
        if decision or impact:
            current[addr] = {"decision": decision, "impact": impact}
        else:
            current.pop(addr, None)
    out = root / "metadata" / "machine_dossier"
    out.mkdir(parents=True, exist_ok=True)
    (out / "decisions.json").write_text(
        json.dumps(current, ensure_ascii=False, indent=1), encoding="utf-8")
    _write_decision_files(root, out)
    return len(current)


def build_decision_rows(root: Path) -> list[list[str]]:
    decisions = load_decisions(root)
    rows = []
    for s in _signals(root):
        addr = s.get("address", "")
        d = decisions.get(addr, {})
        rows.append([
            addr, s.get("name", ""), s.get("type", ""),
            s.get("equipment", ""), s.get("desc", ""),
            s.get("srcmodule", "") or "",
            d.get("decision", ""), d.get("impact", ""),
        ])
    return rows


_DECISION_NOTE = ("ENGINEER DOCUMENT — the two DECISION columns are yours; "
                  "the generator NEVER fills them (they persist in "
                  "decisions.json across regenerations). Every other cell "
                  "is deterministic (RD01 + symbols). DRAFT_UNVERIFIED.")


def _write_decision_files(root: Path, out: Path) -> list[str]:
    rows = build_decision_rows(root)
    written = []
    if _write_xlsx(out / "04_decision_table.xlsx", _DECISION_HEADERS, rows,
                   "Decision table", _DECISION_NOTE):
        written.append("04_decision_table.xlsx")
    _write_md_table(out / "04_decision_table.md",
                    "DECISION TABLE — devices & modernization",
                    _DECISION_NOTE, _DECISION_HEADERS, rows)
    written.append("04_decision_table.md")
    # Old⇄Target page: the structured (KEEP/REPLACE/DROP) reading of the
    # decisions + their cascade. Byproduct of the delta engine — refreshed
    # whenever the decisions change; a delta failure never blocks the save.
    try:
        from decision_cascade import write_delta_file  # type: ignore
        fp = write_delta_file(root)
        if fp is not None:
            written.append(fp.name)
    except Exception:
        pass
    return written


_EQUIP_CLASSES = [
    ("M", "motors / drives"), ("K", "contactors / relays"),
    ("Y", "valves / solenoids"), ("B", "sensors"), ("S", "switches/buttons"),
    ("H", "lamps / horns"), ("F", "protective devices"),
]


def build_plant_summary(root: Path, state: dict,
                        calls: list[tuple[str, str]]) -> str:
    sigs = _signals(root)
    di = sum(1 for s in sigs if s.get("address", "").startswith("%I")
             and "W" not in s.get("address", ""))
    do = sum(1 for s in sigs if s.get("address", "").startswith("%Q")
             and "W" not in s.get("address", ""))
    aio = sum(1 for s in sigs if "W" in s.get("address", ""))
    equips = {s.get("equipment", "").strip() for s in sigs
              if s.get("equipment", "").strip()}
    by_class = []
    for prefix, label in _EQUIP_CLASSES:
        hits = sorted(e for e in equips
                      if re.match(rf"^-?\d*{prefix}\d", e)
                      or e.upper().startswith(prefix))
        if hits:
            by_class.append((label, len(hits), ", ".join(hits[:12])))

    now = datetime.now(timezone.utc).isoformat(timespec="seconds")
    L = [
        "# PLANT SUMMARY — one page (deterministic counts)",
        "",
        f"Generated: {now} · source: RD01 signal list + proven extraction",
        "",
        "> Numbers below are counted, not estimated. The narrative section "
        "at the bottom is intentionally empty — it is the engineer's (or a "
        "clearly-labelled AI draft's) to write.",
        "",
        "## Signals",
        "",
        f"- Digital inputs: **{di}** · digital outputs: **{do}** · "
        f"analog/word: **{aio}** · total rows: **{len(sigs)}**",
        "",
        "## Equipment (from the RD01 Equipment column)",
        "",
    ]
    if by_class:
        for label, n, sample in by_class:
            L.append(f"- {label}: **{n}** ({sample})")
    else:
        L.append("- _Equipment column empty — run the RD01 enrichment first._")
    L += [
        "",
        "## Control structure",
        "",
        f"- Proven step chains: **{len(state['chains'])}** "
        f"({sum(len(c['steps']) for c in state['chains'])} steps)",
        f"- Cross-chain triggers: **{len(state['triggers'])}**",
        f"- Logic networks: **{state['networks']}** "
        f"(proven: {state['networks_proven']}, refused/honest: "
        f"{state['networks'] - state['networks_proven']})",
        f"- Blocks called from OBs: **{len({b for _, b in calls})}**",
        "",
        "## Machine narrative (DRAFT — to be written)",
        "",
        "_(engineer or labelled AI draft — what does this machine make, in "
        "three sentences an operator would recognise?)_",
        "",
    ]
    return "\n".join(L)


def build_ce_rows(state: dict) -> tuple[list[str], list[list[str]]]:
    """Cause rows = distinct atoms of the rail equations; effect columns =
    the chains they gate. X = the atom appears in that chain's enable rail
    equation set (negated or positive — presence is what matters for FAT)."""
    chains = state["chains"]
    headers = ["Cause (signal)", "Symbol"] + [
        _short(c["label"], 24) + " STOP" for c in chains]
    cause_map: dict[str, dict] = {}
    for ci, ch in enumerate(chains):
        rail_addrs = {r["addr"] for r in ch["enable_rails"]}
        for st in ch["steps"]:
            for c in st["conditions"]:
                pass  # step conditions are transitions, not causes
        for r in ch["enable_rails"] + ch["reset_rails"]:
            e = cause_map.setdefault(r["addr"],
                                     {"label": r["label"], "cols": set()})
            e["cols"].add(ci)
    rows = []
    for addr in sorted(cause_map):
        e = cause_map[addr]
        rows.append([addr, e["label"] or "❓"] +
                    ["X" if i in e["cols"] else "—"
                     for i in range(len(chains))])
    return headers, rows


# ---------------------------------------------------------------------------
# customer print copy (PDF via Edge headless — optional, fail-soft)
# ---------------------------------------------------------------------------

def _md_to_html(text: str) -> str:
    """Tiny converter for OUR OWN generated md (headers/bullets/tables)."""
    out, in_table = [], False
    for line in text.splitlines():
        s = line.rstrip()
        if s.startswith("|"):
            cells = [c.strip() for c in s.strip("|").split("|")]
            if all(set(c) <= set("-: ") for c in cells if c):
                continue
            tag = "th" if not in_table else "td"
            if not in_table:
                out.append("<table>")
                in_table = True
            out.append("<tr>" + "".join(
                f"<{tag}>{_esc(c)}</{tag}>" for c in cells) + "</tr>")
            continue
        if in_table:
            out.append("</table>")
            in_table = False
        if s.startswith("# "):
            out.append(f"<h1>{_esc(s[2:])}</h1>")
        elif s.startswith("## "):
            out.append(f"<h2>{_esc(s[3:])}</h2>")
        elif s.startswith("> "):
            out.append(f"<p class='note'>{_esc(s[2:])}</p>")
        elif s.startswith("- "):
            out.append(f"<p class='li'>• {_esc(s[2:])}</p>")
        elif s:
            out.append(f"<p>{_esc(s)}</p>")
    if in_table:
        out.append("</table>")
    return "\n".join(out)


def render_print_html(project_root: Path) -> str | None:
    """One printable HTML: every dossier page in order, for the customer
    PDF. Returns None when no dossier exists yet."""
    dossier = Path(project_root) / "metadata" / "machine_dossier"
    if not dossier.is_dir():
        return None
    theme = load_theme()
    parts = [
        "<!doctype html><html><head><meta charset='utf-8'>",
        f"<title>Machine Dossier — {_esc(Path(project_root).name)}</title>",
        "<style>",
        f"body{{font-family:{theme['font_family']};color:{theme['ink']};"
        "margin:24px}",
        "section{page-break-after:always}",
        "svg{max-width:100%;height:auto}",
        "table{border-collapse:collapse;font-size:10.5px;margin:8px 0}",
        "th,td{border:1px solid #b8c2c8;padding:3px 7px;text-align:left}",
        f"th{{background:{theme['action_fill']}}}",
        f".note{{color:{theme['global_band_text']};font-style:italic}}",
        ".li{margin:2px 0 2px 12px}",
        "h1{font-size:20px}h2{font-size:15px}p{font-size:11.5px;margin:3px 0}",
        "</style></head><body>",
    ]
    order = sorted(dossier.glob("*.svg")) + [
        dossier / "04_decision_table.md",
        dossier / "05_plant_summary.md",
        dossier / "06_ce_matrix.md",
    ]
    added = 0
    for fp in order:
        if not fp.is_file():
            continue
        parts.append("<section>")
        if fp.suffix == ".svg":
            parts.append(fp.read_text(encoding="utf-8"))
        else:
            parts.append(_md_to_html(fp.read_text(encoding="utf-8")))
        parts.append("</section>")
        added += 1
    parts.append("</body></html>")
    return "\n".join(parts) if added else None


_EDGE_CANDIDATES = [
    r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
    r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
]


def export_dossier_pdf(project_root: Path,
                       out_dir: Path | None = None) -> Path | None:
    """Print the dossier to ONE PDF via Edge headless. Fail-soft: returns
    None (never raises) when Edge or the dossier is unavailable.

    The PDF is a HANDOVER artifact (user decision 2026-07-06): it belongs
    in _delivery/, not in the working metadata/machine_dossier/ folder —
    like FAT/SAT it appears when everything is done, at customer handover."""
    import subprocess
    import tempfile
    html = render_print_html(project_root)
    if html is None:
        return None
    edge = next((Path(p) for p in _EDGE_CANDIDATES if Path(p).is_file()), None)
    if edge is None:
        return None
    target = Path(out_dir) if out_dir else (Path(project_root) / "_delivery")
    target.mkdir(parents=True, exist_ok=True)
    out = target / "machine_dossier.pdf"
    try:
        with tempfile.TemporaryDirectory() as td:
            src = Path(td) / "dossier_print.html"
            src.write_text(html, encoding="utf-8")
            subprocess.run(
                [str(edge), "--headless", "--disable-gpu",
                 f"--print-to-pdf={out}", "--no-pdf-header-footer",
                 src.as_uri()],
                capture_output=True, timeout=120, check=False)
        return out if out.is_file() and out.stat().st_size > 0 else None
    except Exception:
        return None


# ---------------------------------------------------------------------------
# main entry
# ---------------------------------------------------------------------------

def generate_machine_dossier(project_root: Path) -> DossierSummary:
    root = Path(project_root)
    theme = load_theme()
    machine = root.name
    out = root / "metadata" / "machine_dossier"
    out.mkdir(parents=True, exist_ok=True)
    summ = DossierSummary(out_dir=out)

    state = build_state_table(root)
    legacy = root / "_raw" / "legacy_code"
    nets = extract_project_logic(root)
    names = load_symbols(legacy)
    calls = parse_block_calls(legacy)
    roles = _block_roles(nets, state, names)

    summ.chains = len(state["chains"])
    summ.steps = sum(len(c["steps"]) for c in state["chains"])
    summ.blocks = len({b for _, b in calls})

    (out / "state_table.json").write_text(
        json.dumps(state, ensure_ascii=False, indent=1), encoding="utf-8")
    summ.files.append("state_table.json")

    (out / "01_operator_flow.svg").write_text(
        render_operator_flow(state, machine, theme), encoding="utf-8")
    summ.files.append("01_operator_flow.svg")

    (out / "02_block_structure.svg").write_text(
        render_block_structure(calls, roles, machine, theme),
        encoding="utf-8")
    summ.files.append("02_block_structure.svg")

    for i, chain in enumerate(state["chains"], 1):
        fp = out / f"03_grafcet_{i}.svg"
        fp.write_text(render_grafcet(chain, machine, theme), encoding="utf-8")
        summ.files.append(fp.name)
    if not state["chains"]:
        summ.warnings.append("no proven step chain — GRAFCET pages skipped")

    summ.signals = len(_signals(root))
    written = _write_decision_files(root, out)
    if "04_decision_table.xlsx" not in written:
        summ.warnings.append("openpyxl unavailable — xlsx skipped")
    summ.files.extend(written)

    (out / "05_plant_summary.md").write_text(
        build_plant_summary(root, state, calls), encoding="utf-8")
    summ.files.append("05_plant_summary.md")

    ce_note = ("Cause→effect from the proven rail equations. X = cause gates "
               "that chain. FAT lines should be tested against THIS table. "
               "DRAFT_UNVERIFIED — engineer confirms.")
    ce_headers, ce_rows = build_ce_rows(state)
    if _write_xlsx(out / "06_ce_matrix.xlsx", ce_headers, ce_rows,
                   "C&E matrix", ce_note):
        summ.files.append("06_ce_matrix.xlsx")
    _write_md_table(out / "06_ce_matrix.md", "C&E MATRIX — interlock rails",
                    ce_note, ce_headers, ce_rows)
    summ.files.append("06_ce_matrix.md")

    return summ


def main() -> int:
    import argparse
    ap = argparse.ArgumentParser(description="Machine Dossier generator")
    ap.add_argument("project_root", type=Path)
    args = ap.parse_args()
    s = generate_machine_dossier(args.project_root)
    print(f"chains: {s.chains} · steps: {s.steps} · blocks: {s.blocks} · "
          f"signal rows: {s.signals}")
    for f in s.files:
        print(f"  wrote {f}")
    for w in s.warnings:
        print(f"  WARN {w}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

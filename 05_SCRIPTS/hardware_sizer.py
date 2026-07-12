#!/usr/bin/env python3
"""
hardware_sizer.py — Hardware Sizing Engine (Phase 24)

Inputs:
  - RD01_IO_List.md     -> channel counts for DI/DQ/AI/AO
  - PROJECT_STATE.json  -> target_platform / target_cpu
  - 20% reserve margin  -> recommended module quantities

Outputs:
  - _output/HW_proposed_BOM.xlsx  -> Excel report
  - _output/HW_proposed_BOM.md    -> AI brief attachment

CLI:
  python hardware_sizer.py --project PROJECT_PATH [--reserve 20]
"""

from __future__ import annotations

import json
import math
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

# R-S-1: merkezi F-CPU tespit yardımcısı
try:
    from workbench.core.safety_utils import is_f_cpu as _is_f_cpu
except ImportError:  # pragma: no cover — fallback if package not on path yet
    def _is_f_cpu(cpu_model):  # type: ignore[misc]
        import re as _re
        return bool(_re.search(r"(?i)(\bSF\b|^SF|\bTF\b|\dF[-\s/]|\dF$|\dTF[-\s/]|\dTF$|[-\s]F[-\s]|[-\s]F$)", cpu_model or ""))


FACTORY_ROOT = Path(__file__).resolve().parent.parent


# -- Module Catalog (built-in — extensible with library MDs) ------------------

@dataclass
class ModuleSpec:
    name: str
    io_type: str           # DI / DQ / AI / AO / PULSE / SAFE_DI / SAFE_DQ
    channels: int
    part_no: str
    platform: str          # S7_1500 / S7_1200 / S7_300 / S7_400 / ALL
    notes: str = ""
    unit_price_eur: float = 0.0   # optional price (0 = unknown)


# ET 200SP modules (standard for S7-1500)
ET200SP_MODULES: list[ModuleSpec] = [
    ModuleSpec("ET200SP DI 16×24VDC HF",       "DI",      16, "6ES7131-6BH01-0BA0", "S7_1500",
               "Fast input (0.05 ms), diagnostic"),
    ModuleSpec("ET200SP DI 8×24VDC HF",         "DI",       8, "6ES7131-6BF01-0BA0", "S7_1500",
               "For compact applications"),
    ModuleSpec("ET200SP DQ 16×24VDC/0.5A HF",   "DQ",      16, "6ES7132-6BH01-0BA0", "S7_1500",
               "0.5A/channel, max 4A/group, short-circuit protected"),
    ModuleSpec("ET200SP DQ 8×24VDC/0.5A HF",    "DQ",       8, "6ES7132-6BF01-0BA0", "S7_1500",
               "Compact application"),
    ModuleSpec("ET200SP AI 4×U/I/RTD ST",        "AI",       4, "6ES7134-6JD00-0CA1", "S7_1500",
               "4-20mA / 0-10V / PT100 / TC — 16 bit"),
    ModuleSpec("ET200SP AI 8×U/I ST",            "AI",       8, "6ES7134-6FF00-0AA1", "S7_1500",
               "U/I only — no RTD"),
    ModuleSpec("ET200SP AO 4×U/I ST",            "AO",       4, "6ES7135-6GB00-0CA1", "S7_1500",
               "0-20mA / 4-20mA / 0-10V / ±10V — 16 bit"),
    ModuleSpec("ET200SP AO 2×U/I ST",            "AO",       2, "6ES7135-6FB00-0CA1", "S7_1500",
               "Compact analog output"),
    # F-modules (Safety)
    ModuleSpec("ET200SP F-DI 8×24VDC HF",        "SAFE_DI",  8, "6ES7136-6BA00-0CA0", "S7_1500",
               "SIL2/PLd — PROFIsafe"),
    ModuleSpec("ET200SP F-DQ 4×24VDC/2A PP HF",  "SAFE_DQ",  4, "6ES7136-6DC00-0CA0", "S7_1500",
               "SIL2/PLd — PROFIsafe, 2A/channel"),
]

# S7-1200 onboard + expansion
S7_1200_MODULES: list[ModuleSpec] = [
    ModuleSpec("SM 1221 DI 16×24VDC",     "DI",  16, "6ES7221-1BH32-0XB0", "S7_1200"),
    ModuleSpec("SM 1221 DI 8×24VDC",      "DI",   8, "6ES7221-1BF32-0XB0", "S7_1200"),
    ModuleSpec("SM 1222 DQ 16×24VDC",     "DQ",  16, "6ES7222-1BH32-0XB0", "S7_1200"),
    ModuleSpec("SM 1222 DQ 8×24VDC",      "DQ",   8, "6ES7222-1BF32-0XB0", "S7_1200"),
    ModuleSpec("SM 1231 AI 4×±10V",       "AI",   4, "6ES7231-4HD32-0XB0", "S7_1200"),
    ModuleSpec("SM 1232 AO 4×±10V/20mA",  "AO",   4, "6ES7232-4HD32-0XB0", "S7_1200"),
]

# S7-300 classic
S7_300_MODULES: list[ModuleSpec] = [
    ModuleSpec("SM323 DI16/DQ16×24V",      "DI",  16, "6ES7323-1BL00-0AA0", "S7_300",
               "Combined DI+DQ module"),
    ModuleSpec("SM321 DI 16×24VDC",        "DI",  16, "6ES7321-1BH02-0AA0", "S7_300"),
    ModuleSpec("SM322 DQ 16×24VDC/0.5A",   "DQ",  16, "6ES7322-1BH01-0AA0", "S7_300"),
    ModuleSpec("SM331 AI 8×12 bit",         "AI",   8, "6ES7331-7KF02-0AB0", "S7_300"),
    ModuleSpec("SM332 AO 4×12 bit",         "AO",   4, "6ES7332-5HD01-0AB0", "S7_300"),
]

PLATFORM_MODULE_MAP: dict[str, list[ModuleSpec]] = {
    "S7_1500": ET200SP_MODULES,
    "S7_1200": S7_1200_MODULES,
    "S7_300":  S7_300_MODULES,
    "S7_400":  S7_300_MODULES,  # S7-400 usually uses the same ET200M / SM
    "AB_LOGIX": [],  # ControlLogix — custom
    "BECKHOFF": [],  # TwinCAT — varies by EtherCAT terminals
    "CODESYS": [],
}

# Head-station catalog (for S7-1500 + ET200SP)
HEAD_STATION_MAP: dict[str, dict] = {
    "S7_1500": {
        "name": "ET200SP IM 155-6 PN ST",
        "part_no": "6ES7155-6AU01-0BN0",
        "notes": "PROFINET head station — max 32 modules",
        "max_modules": 32,
    },
    "S7_1200": {
        "name": "S7-1200 (onboard I/O + SM expansion)",
        "part_no": "—",
        "notes": "Max 8 SM expansion modules",
        "max_modules": 8,
    },
    "S7_300": {
        "name": "S7-300 Rack (UR1/UR2)",
        "part_no": "6ES7390-1AE80-0AA0",
        "notes": "Max 8 modules/rack; for extra racks use IM365/IM360+361",
        "max_modules": 8,
    },
}


# -- IO Count -----------------------------------------------------------------

@dataclass
class IOCount:
    di: int = 0
    dq: int = 0
    ai: int = 0
    ao: int = 0
    pulse: int = 0
    safe_di: int = 0
    safe_dq: int = 0
    source: str = "?"

    @property
    def total(self) -> int:
        return self.di + self.dq + self.ai + self.ao + self.pulse + self.safe_di + self.safe_dq


def _is_safe_type(type_upper: str) -> bool:
    """Return True if the Type value already identifies a safety-channel explicitly."""
    return type_upper in ("F-DI", "FDI", "SAFE_DI", "F-DQ", "FDQ", "SAFE_DQ",
                          "F_DI", "F_DQ", "SAFE")


def _is_safe_di_type(type_upper: str) -> bool:
    return type_upper in ("F-DI", "FDI", "SAFE_DI", "F_DI")


def _is_safe_dq_type(type_upper: str) -> bool:
    return type_upper in ("F-DQ", "FDQ", "SAFE_DQ", "F_DQ")


def count_from_rd01(rd01_path: Path) -> IOCount:
    """Count IO types from RD01_IO_List.md.

    N-C2 fix: SafetyRelated=Y column is now read.  A row whose Type is a
    standard signal (DI/DQ/…) but whose SafetyRelated column is 'Y' (case-
    insensitive) is promoted to safe_di / safe_dq instead of di / dq — so the
    F-CPU sizing check receives the correct counts.

    Column discovery: the header row is parsed first to locate 'Type',
    'Direction', and 'SafetyRelated' by name.  Rows without a recognisable
    header fall back to the legacy positional heuristic (cols[1:3]).
    """
    count = IOCount(source=str(rd01_path))
    if not rd01_path.exists():
        return count

    content = rd01_path.read_text(encoding="utf-8")

    table_pattern = re.compile(r"^\s*\|(.+)\|", re.MULTILINE)

    # Header-column indices (populated when we find the header row)
    idx_type: Optional[int] = None
    idx_dir:  Optional[int] = None
    idx_safe: Optional[int] = None
    header_found = False

    _HEADER_TAGS = ("TAG", "SEMBOL", "AD", "NO", "ID", "NAME", "—", "-")
    _HEADER_SKIP = re.compile(r"^[-:| ]+$")  # separator lines like |---|---|

    for match in table_pattern.finditer(content):
        cols = [c.strip() for c in match.group(1).split("|")]
        if len(cols) < 2:
            continue

        # Detect separator row (|---|---|)
        joined = "".join(cols)
        if _HEADER_SKIP.match(joined):
            continue

        # Detect header row — first column is a known header keyword
        tag_col_u = cols[0].upper()
        if not header_found and tag_col_u in _HEADER_TAGS:
            # Map column names to indices (normalise: lower, strip spaces/underscore)
            def _norm(s: str) -> str:
                return s.lower().replace(" ", "").replace("_", "")
            header_found = True
            for ci, hdr in enumerate(cols):
                hn = _norm(hdr)
                if hn in ("type", "tip"):
                    idx_type = ci
                elif hn in ("direction", "yön", "yon"):
                    idx_dir = ci
                elif hn in ("safetyrelated", "safety_related", "güvenlik", "guvenlik"):
                    idx_safe = ci
            continue

        # Data row
        if tag_col_u in _HEADER_TAGS:
            continue  # extra header rows in multi-table documents

        # Resolve IO type
        if idx_type is not None and idx_type < len(cols):
            io_type_col = cols[idx_type].upper()
        else:
            # Legacy fallback: scan cols[1:3] for a known type keyword
            io_type_col = ""
            for col in cols[1:3]:
                col_upper = col.upper()
                if col_upper in ("DI", "DO", "DQ", "AI", "AO", "ANALOG_IN", "ANALOG_OUT",
                                 "BOOL", "REAL", "PULSE", "F-DI", "F-DQ", "FDI", "FDQ",
                                 "SAFE", "SAFETY"):
                    io_type_col = col_upper
                    break
            if not io_type_col:
                continue

        # Resolve SafetyRelated flag
        safety_y = False
        if idx_safe is not None and idx_safe < len(cols):
            safety_y = cols[idx_safe].strip().upper() == "Y"

        # Resolve Direction for ambiguous promotions (DI vs DQ when Type=BOOL etc.)
        direction_u = ""
        if idx_dir is not None and idx_dir < len(cols):
            direction_u = cols[idx_dir].strip().upper()

        # Classify
        if _is_safe_di_type(io_type_col):
            count.safe_di += 1
        elif _is_safe_dq_type(io_type_col):
            count.safe_dq += 1
        elif io_type_col == "SAFE" or (safety_y and _is_safe_type(io_type_col)):
            # 'SAFE' column value without specific subtype — direction-based split
            if direction_u in ("DQ", "DO", "OUTPUT"):
                count.safe_dq += 1
            else:
                count.safe_di += 1
        elif io_type_col in ("DI", "BOOL") and safety_y:
            # N-C2: standard DI but SafetyRelated=Y → promote to safe_di
            count.safe_di += 1
        elif io_type_col in ("DQ", "DO") and safety_y:
            # N-C2: standard DQ but SafetyRelated=Y → promote to safe_dq
            count.safe_dq += 1
        elif io_type_col in ("DI", "BOOL"):
            count.di += 1
        elif io_type_col in ("DQ", "DO"):
            count.dq += 1
        elif io_type_col in ("AI", "ANALOG_IN", "REAL"):
            count.ai += 1
        elif io_type_col in ("AO", "ANALOG_OUT"):
            count.ao += 1
        elif io_type_col == "PULSE":
            count.pulse += 1

    return count


def count_from_hw_config(hw_config_path: Path) -> Optional[IOCount]:
    """Get the IO count from hardware_config.xlsx (more reliable if parsed).

    S-1 / B-L8: SAFE_DI / SAFE_DQ entries in the HW config file are now
    counted into the dedicated safe_di / safe_dq fields instead of being
    silently dropped (the previous elif-chain had no branch for them, so they
    fell through without incrementing any counter — producing an honest-looking
    IOCount that quietly under-reported safety channels).
    """
    try:
        from hw_config_parser import parse_hw_config
        result = parse_hw_config(hw_config_path)
        if result.errors:
            return None
        count = IOCount(source=str(hw_config_path))
        for entry in result.entries:
            io_t = entry.io_type.upper()
            if _is_safe_di_type(io_t):
                # S-1: SAFE IO requires F-module — not sized automatically via
                # this path; counted here so the caller can route to size_modules
                # which enforces the F-CPU requirement (C-A3, fail-closed).
                count.safe_di += 1
            elif _is_safe_dq_type(io_t):
                count.safe_dq += 1
            elif io_t == "DI":
                count.di += 1
            elif io_t in ("DQ", "DO"):
                count.dq += 1
            elif io_t == "AI":
                count.ai += 1
            elif io_t == "AO":
                count.ao += 1
            # else: unknown type — silently skip (same as before for non-SAFE unknowns)
        return count
    except Exception:
        return None


# -- Sizing engine ------------------------------------------------------------

@dataclass
class ModuleRecommendation:
    io_type: str
    raw_count: int          # raw count from RD01
    with_reserve: int       # count with xx% reserve
    module: ModuleSpec
    module_count: int       # how many modules are needed
    total_channels: int     # total channels of the modules
    spare_channels: int     # spare channels
    note: str = ""


@dataclass
class SizerResult:
    io_count: IOCount
    reserve_pct: int
    platform: str
    cpu: str
    recommendations: list[ModuleRecommendation] = field(default_factory=list)
    head_station: Optional[dict] = None
    total_modules: int = 0
    warnings: list[str] = field(default_factory=list)
    # C-A3: hard errors. When non-empty, downstream code (BOM xlsx, customer
    # report, AI brief) MUST refuse to produce a deliverable that pretends the
    # sizing is valid. SAFE_* channels sized against a non-F-CPU populate this.
    errors: list[str] = field(default_factory=list)
    safety_misconfigured: bool = False


class SafetyMisconfigurationError(ValueError):
    """C-A3: raised when SAFE_DI/SAFE_DQ channels are requested on a CPU that
    is not an F-CPU. The previous code logged a warning and silently mapped
    safety IO to standard modules — producing a BOM and a customer report
    that quietly violates SIL/PLr requirements. We now fail closed."""


def _best_module(modules: list[ModuleSpec], io_type: str, prefer_large: bool = True) -> Optional[ModuleSpec]:
    """Pick the most suitable module for the requested IO type."""
    candidates = [m for m in modules if m.io_type == io_type]
    if not candidates:
        return None
    # Large first, then small (or vice versa)
    candidates.sort(key=lambda m: m.channels, reverse=prefer_large)
    return candidates[0]


def size_modules(
    io_count: IOCount,
    platform: str,
    cpu: str = "",
    reserve_pct: int = 20,
    *,
    strict_safety: bool = True,
) -> SizerResult:
    """Produce a module recommendation based on the IO counts.

    C-A3: ``strict_safety`` (default True) means SAFE_DI/SAFE_DQ channels on a
    non-F-CPU raise :class:`SafetyMisconfigurationError` and DO NOT silently
    fall back to standard modules. Set ``strict_safety=False`` to keep the
    legacy "warn + downgrade" behaviour, which should only ever be used by
    intentionally-unsafe diagnostic tooling — never by the BOM or the
    customer-facing report.
    """
    result = SizerResult(
        io_count=io_count,
        reserve_pct=reserve_pct,
        platform=platform,
        cpu=cpu,
    )

    modules = PLATFORM_MODULE_MAP.get(platform, [])
    if not modules:
        result.warnings.append(
            f"No built-in module catalog for '{platform}'. "
            "Select the modules manually."
        )
        return result

    # IO types to size
    sizing_map = [
        ("DI",      io_count.di,      False),  # (type, count, safety)
        ("DQ",      io_count.dq,      False),
        ("AI",      io_count.ai,      False),
        ("AO",      io_count.ao,      False),
        ("SAFE_DI", io_count.safe_di, True),
        ("SAFE_DQ", io_count.safe_dq, True),
    ]

    is_safety_cpu = _is_f_cpu(cpu)  # R-S-1: merkezi F-CPU tespiti
    total_modules = 0

    # C-A3: pre-scan for safety/CPU mismatch and either raise (strict) or
    # record an error + skip the SAFE_* rows. We never silently downgrade.
    safety_channel_total = (
        (io_count.safe_di or 0) + (io_count.safe_dq or 0)
        if any(io_count.__dict__.get(k) for k in ("safe_di", "safe_dq")) else 0
    )
    if safety_channel_total > 0 and not is_safety_cpu:
        msg = (
            f"RD01 declares {safety_channel_total} SAFE_* channel(s) but "
            f"target CPU '{cpu or '(unset)'}' is not an F-CPU. "
            "Refusing to size safety IO with standard modules — choose an "
            "F-CPU (e.g. 'CPU 1515F-2 PN') or remove the safety channels "
            "from RD01. (C-A3, fail-closed)"
        )
        if strict_safety:
            raise SafetyMisconfigurationError(msg)
        result.errors.append(msg)
        result.safety_misconfigured = True

    for io_type, raw_count, is_safety in sizing_map:
        if raw_count == 0:
            continue
        if is_safety and not is_safety_cpu:
            # Already recorded above; SAFE_* rows are dropped to keep the BOM
            # from quietly listing standard modules in place of F-modules.
            continue

        with_reserve = math.ceil(raw_count * (1 + reserve_pct / 100))
        mod = _best_module(modules, io_type)
        if mod is None:
            # S-1 / B-L8: SAFE IO requires an F-module from the platform catalog.
            # If the catalog has no F-module for this platform (e.g. S7_1200 /
            # S7_300 which do not carry PROFIsafe F-DI/F-DQ in this built-in
            # catalog), we emit an explicit, actionable warning rather than
            # silently skipping the row and producing an incomplete BOM.
            if io_type in ("SAFE_DI", "SAFE_DQ"):
                result.warnings.append(
                    f"SAFE IO requires F-module — not sized automatically. "
                    f"{raw_count} {io_type} channel(s) detected but no F-module "
                    f"found in the '{platform}' catalog. "
                    "Add the appropriate PROFIsafe F-DI/F-DQ module manually."
                )
            else:
                result.warnings.append(f"No suitable module found for '{io_type}'.")
            continue

        mod_count = math.ceil(with_reserve / mod.channels)
        total_ch  = mod_count * mod.channels
        spare_ch  = total_ch - with_reserve

        rec = ModuleRecommendation(
            io_type=io_type,
            raw_count=raw_count,
            with_reserve=with_reserve,
            module=mod,
            module_count=mod_count,
            total_channels=total_ch,
            spare_channels=spare_ch,
            note=f"+{reserve_pct}% reserve ({raw_count} -> {with_reserve} channels)",
        )
        result.recommendations.append(rec)
        total_modules += mod_count

    result.total_modules = total_modules

    # Head-station recommendation
    head_info = HEAD_STATION_MAP.get(platform)
    if head_info:
        result.head_station = head_info.copy()
        if total_modules > head_info["max_modules"]:
            extra = total_modules - head_info["max_modules"]
            result.warnings.append(
                f"Total {total_modules} modules exceeds a single IM's max of {head_info['max_modules']} modules. "
                f"An extra {math.ceil(extra / head_info['max_modules'])} IM station(s) may be needed."
            )

    # Pulse IO warning
    if io_count.pulse > 0:
        result.warnings.append(
            f"{io_count.pulse} PULSE inputs detected. "
            "A high-speed counter/encoder module may be needed — select manually."
        )

    return result


# -- Excel output -------------------------------------------------------------

def generate_sizer_xlsx(result: SizerResult, output_path: Path) -> bool:
    # C-A3: refuse to produce a BOM when the sizing has hard errors (e.g.
    # safety/CPU mismatch). The customer must not receive an authoritative-
    # looking spreadsheet whose underlying calculation is unsafe.
    if result.errors:
        return False
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed: pip install openpyxl", file=sys.stderr)
        return False

    wb = openpyxl.Workbook()

    # -- Sheet 1: Module Recommendation -----------------------------------
    ws = wb.active
    ws.title = "Proposed_BOM"

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    grn_fill = PatternFill("solid", fgColor="D6F5DC")
    grn_font = Font(color="1A5C2A", bold=True, name="Calibri", size=10)
    wrn_fill = PatternFill("solid", fgColor="FFF2CC")
    data_font = Font(name="Calibri", size=10)
    even_fill = PatternFill("solid", fgColor="EBF5FB")
    border = Border(bottom=Side(style="thin"), right=Side(style="thin"))

    headers = [
        ("IO Type",          10), ("Raw Count", 10), (f"+{result.reserve_pct}% Reserve", 14),
        ("Recommended Module", 36), ("Order No",  20), ("Qty",   7),
        ("Total Channels",   14), ("Spare Channels",  14), ("Notes", 36),
    ]

    for col_idx, (hdr, width) in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        c = ws.cell(row=1, column=col_idx, value=hdr)
        c.font, c.fill = hdr_font, hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    ws.row_dimensions[1].height = 18

    for r_off, rec in enumerate(result.recommendations, start=2):
        fill = even_fill if r_off % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        row_data = [
            rec.io_type, rec.raw_count, rec.with_reserve,
            rec.module.name, rec.module.part_no, rec.module_count,
            rec.total_channels, rec.spare_channels, rec.note,
        ]
        for col_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=r_off, column=col_idx, value=val)
            c.font, c.fill, c.border = data_font, fill, border
            c.alignment = Alignment(horizontal="left", vertical="center")

    # Total row
    total_row = len(result.recommendations) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, name="Calibri", size=10)
    ws.cell(row=total_row, column=6, value=result.total_modules).font = Font(bold=True, name="Calibri", size=10)

    # Head station
    if result.head_station:
        ws.cell(row=total_row + 2, column=1, value="Head Station:").font = Font(bold=True)
        ws.cell(row=total_row + 2, column=2, value=result.head_station["name"])
        ws.cell(row=total_row + 2, column=5, value=result.head_station["part_no"])
        ws.cell(row=total_row + 2, column=6, value=1)
        ws.cell(row=total_row + 2, column=9, value=result.head_station["notes"])

    # Warnings
    if result.warnings:
        ws.cell(row=total_row + 4, column=1, value="WARNINGS:").font = Font(bold=True, color="CC5500")
        for w_idx, warn in enumerate(result.warnings):
            c = ws.cell(row=total_row + 5 + w_idx, column=1, value=warn)
            c.fill = wrn_fill
            c.font = Font(name="Calibri", size=10, color="7D3C00")

    # -- Sheet 2: IO Summary ----------------------------------------------
    ws2 = wb.create_sheet("IO_Summary")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 16

    ws2.cell(row=1, column=1, value="IO Summary — Source").font = Font(bold=True, size=12)
    ws2.cell(row=1, column=2, value=result.io_count.source).font = Font(size=10, italic=True)

    summary_rows = [
        ("Digital Inputs (DI):", result.io_count.di),
        ("Digital Outputs (DQ):", result.io_count.dq),
        ("Analog Inputs (AI):", result.io_count.ai),
        ("Analog Outputs (AO):", result.io_count.ao),
        ("Pulse/Encoder (PULSE):", result.io_count.pulse),
        ("Safety DI (F-DI):", result.io_count.safe_di),
        ("Safety DQ (F-DQ):", result.io_count.safe_dq),
        ("TOTAL:", result.io_count.total),
    ]
    for r_idx, (label, val) in enumerate(summary_rows, start=3):
        ws2.cell(row=r_idx, column=1, value=label).font = Font(bold=(label == "TOTAL:"), size=10)
        ws2.cell(row=r_idx, column=2, value=val).font = Font(bold=(label == "TOTAL:"), size=10)
        if label == "TOTAL:":
            ws2.cell(row=r_idx, column=1).fill = grn_fill
            ws2.cell(row=r_idx, column=2).fill = grn_fill

    ws2.cell(row=13, column=1, value=f"Reserve Margin:").font = Font(size=10)
    ws2.cell(row=13, column=2, value=f"{result.reserve_pct}%").font = Font(bold=True, size=10)
    ws2.cell(row=14, column=1, value="Target Platform:").font = Font(size=10)
    ws2.cell(row=14, column=2, value=result.platform).font = Font(size=10)
    ws2.cell(row=15, column=1, value="CPU:").font = Font(size=10)
    ws2.cell(row=15, column=2, value=result.cpu).font = Font(size=10)

    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws2.cell(row=17, column=1, value=f"Generated: {ts}").font = Font(size=9, italic=True, color="888888")
    ws2.cell(row=18, column=1, value="AUTOMATION_FACTORY hardware_sizer.py").font = Font(size=9, italic=True, color="888888")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return True


# -- MD output ----------------------------------------------------------------

def generate_sizer_md(result: SizerResult, output_path: Path) -> bool:
    # C-A3: refuse to write the AI brief MD when sizing has hard errors.
    if result.errors:
        return False
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    lines = []

    lines.append("# HW03 — Proposed Hardware BOM (Auto Sizing)")
    lines.append("")
    lines.append("```yaml")
    lines.append("document_id: HW03")
    lines.append("title: Proposed Hardware BOM")
    lines.append(f"platform: {result.platform}")
    lines.append(f"cpu: {result.cpu}")
    lines.append(f"reserve_margin: {result.reserve_pct}%")
    lines.append(f"updated: {ts}")
    lines.append(f"total_modules: {result.total_modules}")
    lines.append("```")
    lines.append("")
    lines.append("> **Note:** This recommendation is the result of an automatic calculation.")
    lines.append("> The final hardware list must be approved by the project engineer.")
    lines.append("")

    # IO summary
    lines.append("## IO Channel Count")
    lines.append("")
    lines.append(f"| IO Type | Raw Count | +{result.reserve_pct}% Reserve |")
    lines.append("|---------|-----------|---------------|")
    io = result.io_count
    if io.di > 0:      lines.append(f"| Digital Input (DI)  | {io.di}  | {math.ceil(io.di  * (1 + result.reserve_pct/100))} |")
    if io.dq > 0:      lines.append(f"| Digital Output (DQ) | {io.dq}  | {math.ceil(io.dq  * (1 + result.reserve_pct/100))} |")
    if io.ai > 0:      lines.append(f"| Analog Input (AI)   | {io.ai}  | {math.ceil(io.ai  * (1 + result.reserve_pct/100))} |")
    if io.ao > 0:      lines.append(f"| Analog Output (AO)  | {io.ao}  | {math.ceil(io.ao  * (1 + result.reserve_pct/100))} |")
    if io.safe_di > 0: lines.append(f"| Safety DI (F-DI)    | {io.safe_di} | {math.ceil(io.safe_di * (1 + result.reserve_pct/100))} |")
    if io.safe_dq > 0: lines.append(f"| Safety DQ (F-DQ)    | {io.safe_dq} | {math.ceil(io.safe_dq * (1 + result.reserve_pct/100))} |")
    lines.append(f"| **TOTAL**           | **{io.total}** | — |")
    lines.append("")

    # Module recommendations
    lines.append("## Recommended Modules")
    lines.append("")
    lines.append("| IO Type | Module | Order No | Qty | Total Ch | Spare Ch |")
    lines.append("|---------|--------|----------|-----|----------|----------|")
    for rec in result.recommendations:
        lines.append(
            f"| {rec.io_type} | {rec.module.name} | `{rec.module.part_no}` "
            f"| **{rec.module_count}** | {rec.total_channels} | {rec.spare_channels} |"
        )

    # Head station
    if result.head_station:
        lines.append("")
        lines.append("### Head Station")
        lines.append(f"| Module | Order No | Qty | Note |")
        lines.append("|--------|----------|-----|------|")
        hs = result.head_station
        lines.append(f"| {hs['name']} | `{hs['part_no']}` | 1 | {hs['notes']} |")

    lines.append("")
    lines.append(f"**Total module count: {result.total_modules}**")
    lines.append("")

    # Warnings
    if result.warnings:
        lines.append("## Warnings")
        lines.append("")
        for w in result.warnings:
            lines.append(f"- {w}")
        lines.append("")

    lines.append("---")
    lines.append(f"*AUTOMATION_FACTORY hardware_sizer.py — {ts}*")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines), encoding="utf-8")
    return True


# -- Main function (for GUI and CLI) ------------------------------------------

def run_sizer(project_path: Path, reserve_pct: int = 20) -> Optional[SizerResult]:
    """Size the project. None = error."""
    state_file = project_path / "PROJECT_STATE.json"
    if not state_file.exists():
        return None

    try:
        state = json.loads(state_file.read_text(encoding="utf-8"))
    except Exception:
        return None

    platform = state.get("target_platform", "S7_1500")
    cpu      = state.get("target_cpu", "")

    # IO count — hardware_config.xlsx first, then RD01
    hw_xlsx = project_path / "_input" / "hardware_config.xlsx"
    io_count = None

    if hw_xlsx.exists():
        io_count = count_from_hw_config(hw_xlsx)

    if io_count is None or io_count.total == 0:
        rd01 = project_path / "metadata" / "RD01_IO_List.md"
        if rd01.exists():
            io_count = count_from_rd01(rd01)
        else:
            io_count = IOCount(source="(source not found)")

    # C-A3: run the sizer in non-strict mode so we get back a SizerResult
    # carrying `errors`/`safety_misconfigured`. Downstream output writers
    # (generate_sizer_xlsx/md) check those and refuse to produce a deliverable.
    # Programmatic callers who want a hard exception should call size_modules
    # directly with the default strict_safety=True.
    return size_modules(io_count, platform, cpu, reserve_pct, strict_safety=False)


# -- CLI ----------------------------------------------------------------------

def main():
    import argparse
    p = argparse.ArgumentParser(description="Hardware Sizer")
    p.add_argument("--project", metavar="PATH", required=True, help="Project path")
    p.add_argument("--reserve", type=int, default=20, help="Reserve margin %% (default: 20)")
    p.add_argument("--no-xlsx", action="store_true", help="Do not produce Excel output")
    p.add_argument("--no-md", action="store_true", help="Do not produce MD output")
    args = p.parse_args()

    project_path = Path(args.project)
    result = run_sizer(project_path, args.reserve)
    if result is None:
        print(f"Could not read project: {project_path}")
        return

    io = result.io_count
    print(f"IO Count ({io.source}):")
    print(f"   DI={io.di}  DQ={io.dq}  AI={io.ai}  AO={io.ao}  Total={io.total}")
    print(f"   +{result.reserve_pct}% reserve margin")
    print()

    for rec in result.recommendations:
        print(f"   {rec.io_type}: {rec.module_count}x {rec.module.name} "
              f"({rec.module.part_no}) — {rec.total_channels} channels ({rec.spare_channels} spare)")
    if result.head_station:
        print(f"   Head Station: {result.head_station['name']} ({result.head_station['part_no']})")
    print(f"   Total modules: {result.total_modules}")

    if result.warnings:
        print()
        for w in result.warnings:
            print(f"   WARN: {w}")

    if result.errors:
        print()
        for e in result.errors:
            print(f"   ERROR: {e}")
        print("\nBOM and AI brief NOT written — fix the errors above first.")
        sys.exit(2)

    if not args.no_xlsx:
        xlsx_out = project_path / "_output" / "HW_proposed_BOM.xlsx"
        ok = generate_sizer_xlsx(result, xlsx_out)
        if ok:
            print(f"\nExcel: {xlsx_out}")

    if not args.no_md:
        md_out = project_path / "_output" / "HW_proposed_BOM.md"
        ok = generate_sizer_md(result, md_out)
        if ok:
            print(f"MD: {md_out}")


if __name__ == "__main__":
    main()

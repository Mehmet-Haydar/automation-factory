#!/usr/bin/env python3
"""
tia_tag_export.py — TIA Portal Tag Table XML Export (Phase 29-B)

Generates a TIA Portal-importable XML tag table from the RD01 IO list +
HW03 IEC tags.

Source priority:
  1. metadata/HW03_IEC_Tags.md  (if iec_tag_generator.py has been run)
  2. metadata/RD01_IO_List.md   (parsed directly)

Addresses:
  - If HW03/RD01 has a real address (%I0.0 etc.) -> used as-is
  - Otherwise -> a type-based sequential address is auto-generated (with a warning)

TIA Portal XML format: SW.PlcTagTable (V14-V20 compatible)
Optional XLSX output: for review

CLI:
  python tia_tag_export.py --project PROJECT_PATH [--out FOLDER] [--table TABLE_NAME]
"""

from __future__ import annotations

import json
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional


FACTORY_ROOT = Path(__file__).resolve().parent.parent

# Signal type -> TIA Portal data type
_DATA_TYPE: dict[str, str] = {
    "DI":  "Bool",
    "DQ":  "Bool",
    "DO":  "Bool",   # old notation
    "AI":  "Int",
    "AQ":  "Int",
    "AO":  "Int",    # old notation
    "UNK": "Bool",
}

# Address prefix: for sequential auto-generation
_ADDR_PREFIX: dict[str, str] = {
    "DI": "I",
    "DQ": "Q",
    "DO": "Q",
    "AI": "IW",
    "AQ": "QW",
    "AO": "QW",
    "UNK": "I",
}

# Analog starting word address (ET 200SP typical)
_ANALOG_START_WORD = 256


# -- Data structures ----------------------------------------------------------

@dataclass
class PlcTag:
    name: str           # IEC-compliant tag name
    data_type: str      # Bool, Int, Word, Real …
    address: str        # %I0.0 / %Q0.0 / %IW256 / "" (empty = unassigned)
    comment: str        # Description
    signal_type: str    # DI/DQ/AI/AQ/UNK
    original: str       # original name in RD01


@dataclass
class ExportResult:
    xml_path: Optional[Path] = None
    xlsx_path: Optional[Path] = None
    tags: list[PlcTag] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    auto_addressed: int = 0

    @property
    def ok(self) -> bool:
        return self.xml_path is not None


# -- Address auto-generation --------------------------------------------------

class _AddrGen:
    """Type-based sequential address generator."""

    def __init__(self):
        self._digital: dict[str, int] = {"I": 0, "Q": 0}  # byte counter
        self._digital_bit: dict[str, int] = {"I": 0, "Q": 0}
        self._analog: dict[str, int] = {"IW": _ANALOG_START_WORD, "QW": _ANALOG_START_WORD}

    def next(self, signal_type: str) -> tuple[str, bool]:
        """Return (address_string, is_auto)."""
        prefix = _ADDR_PREFIX.get(signal_type, "I")
        if prefix in ("I", "Q"):
            byte = self._digital[prefix]
            bit  = self._digital_bit[prefix]
            addr = f"%{prefix}{byte}.{bit}"
            self._digital_bit[prefix] += 1
            if self._digital_bit[prefix] > 7:
                self._digital_bit[prefix] = 0
                self._digital[prefix] += 1
        else:  # IW, QW
            word = self._analog[prefix]
            addr = f"%{prefix}{word}"
            self._analog[prefix] += 2
        return addr, True


_ADDR_RE = re.compile(r"%[IQM][WDB]?\d+(\.\d+)?", re.IGNORECASE)


def _normalize_address(raw: str) -> str:
    """Raw address -> TIA Portal format or empty string."""
    if not raw or raw in ("-", "—", "?", "TBD"):
        return ""
    m = _ADDR_RE.search(raw)
    if m:
        return m.group(0).upper()
    return ""


# -- HW03 Parser --------------------------------------------------------------

def _parse_hw03(project_path: Path) -> list[PlcTag]:
    """
    Read the tag list from metadata/HW03_IEC_Tags.md.
    Expected header: | IEC Tag Name | Type | Address | Original Name | Description |
    (Turkish column keywords kept to match legacy HW03 files.)
    """
    hw03 = project_path / "metadata" / "HW03_IEC_Tags.md"
    if not hw03.exists():
        return []

    tags: list[PlcTag] = []
    lines = hw03.read_text(encoding="utf-8", errors="ignore").splitlines()
    in_table = False
    col_idx: dict[str, int] = {}

    for line in lines:
        s = line.strip()
        if not s.startswith("|"):
            if in_table:
                break
            continue
        if re.match(r"^\|[-:\s|]+\|$", s):
            continue

        cells = [c.strip().strip("`") for c in s.split("|")[1:-1]]
        if not cells:
            continue

        if not in_table:
            lower = [c.lower() for c in cells]
            for i, c in enumerate(lower):
                if "iec tag" in c or "tag" in c:
                    col_idx["name"] = i
                elif "tip" in c or "type" in c:
                    col_idx["type"] = i
                elif "adres" in c or "address" in c:
                    col_idx["addr"] = i
                elif "orijinal" in c or "original" in c:
                    col_idx["orig"] = i
                elif "açıklama" in c or "desc" in c or "comment" in c:
                    col_idx["desc"] = i
            if "name" in col_idx:
                in_table = True
            continue

        def _get(key: str) -> str:
            idx = col_idx.get(key, -1)
            return cells[idx] if 0 <= idx < len(cells) else ""

        name     = _get("name")
        sig_type = _get("type").strip().upper()
        addr_raw = _get("addr")
        orig     = _get("orig")
        desc     = _get("desc")

        if not name or name == "-":
            continue

        sig_type = sig_type if sig_type in _DATA_TYPE else "UNK"
        addr = _normalize_address(addr_raw)

        tags.append(PlcTag(
            name=name,
            data_type=_DATA_TYPE.get(sig_type, "Bool"),
            address=addr,
            comment=desc,
            signal_type=sig_type,
            original=orig,
        ))

    return tags


# -- RD01 Fallback Parser -----------------------------------------------------

_NAME_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")


def _parse_rd01_direct(project_path: Path, use_original: bool = False) -> list[PlcTag]:
    """Parse RD01 directly (HW03 absent or program names requested).

    ``use_original=True`` keeps the raw RD01 signal name (MOT_HYD_001_FBM)
    instead of the IEC-prefixed one (DI_MOT_HYD_001_FBM). The program
    assembler binds OB1 against the raw RD01 names, so the tag table sent
    to TIA must carry exactly those — the 2026-06-10 live test compiled
    with 11 "Tag not defined" errors because of the prefix mismatch.
    """
    try:
        from iec_tag_generator import parse_rd01_signals, generate_tags
    except ImportError:
        return []

    signals = parse_rd01_signals(project_path)
    if not signals:
        return []

    result = generate_tags(signals)
    tags = []
    for t in result.tags:
        addr = _normalize_address(t.address)
        name = t.tag_name
        if use_original and t.original_name and _NAME_RE.match(t.original_name):
            name = t.original_name
        tags.append(PlcTag(
            name=name,
            data_type=_DATA_TYPE.get(t.signal_type, "Bool"),
            address=addr,
            comment=t.description,
            signal_type=t.signal_type,
            original=t.original_name,
        ))
    return tags


# -- XML Generator ------------------------------------------------------------

def _load_tia_version(project_path: Path) -> str:
    state_file = project_path / "PROJECT_STATE.json"
    if state_file.exists():
        try:
            state = json.loads(state_file.read_text(encoding="utf-8"))
            ver = state.get("target_tia_version", "")
            if ver:
                return ver.strip()
        except Exception:
            pass
    return "V18"


def _load_comment_culture(project_path: Path) -> str:
    """PROJECT_STATE output_language -> TIA comment culture.

    Openness rejects an import whose comment Culture is not a project
    language, so the culture must follow the project's output language.
    """
    lang = ""
    state_file = project_path / "PROJECT_STATE.json"
    if state_file.exists():
        try:
            state = json.loads(state_file.read_text(encoding="utf-8"))
            lang = str(state.get("output_language", "")).strip().upper()
        except Exception:
            pass
    return {"DE": "de-DE", "EN": "en-US", "TR": "tr-TR"}.get(lang, "en-US")


def _build_xml(
    tags: list[PlcTag],
    table_name: str,
    tia_version: str,
    culture: str = "en-US",
) -> str:
    # Openness V14+ import format: SW.Tags.PlcTagTable, document-unique hex
    # IDs, and tag comments as MultilingualText objects. The old
    # SW.PlcTagTable element with an inline <Comment> attribute is the V13
    # shape — TIA V19 refuses it ("unknown element").
    root = ET.Element("Document")
    eng  = ET.SubElement(root, "Engineering")
    eng.set("version", tia_version)

    _counter = iter(range(0, 1 << 30))

    def _next_id() -> str:
        return format(next(_counter), "X")

    plc_table = ET.SubElement(root, "SW.Tags.PlcTagTable")
    plc_table.set("ID", _next_id())

    attr_list = ET.SubElement(plc_table, "AttributeList")
    ET.SubElement(attr_list, "Name").text = table_name

    obj_list = ET.SubElement(plc_table, "ObjectList")

    for tag in tags:
        plc_tag = ET.SubElement(obj_list, "SW.Tags.PlcTag")
        plc_tag.set("ID", _next_id())
        plc_tag.set("CompositionName", "Tags")

        tag_attrs = ET.SubElement(plc_tag, "AttributeList")
        ET.SubElement(tag_attrs, "DataTypeName").text = tag.data_type
        if tag.address:
            ET.SubElement(tag_attrs, "LogicalAddress").text = tag.address
        ET.SubElement(tag_attrs, "Name").text = tag.name

        if tag.comment:
            tag_objs = ET.SubElement(plc_tag, "ObjectList")
            ml = ET.SubElement(tag_objs, "MultilingualText")
            ml.set("ID", _next_id())
            ml.set("CompositionName", "Comment")
            ml_objs = ET.SubElement(ml, "ObjectList")
            item = ET.SubElement(ml_objs, "MultilingualTextItem")
            item.set("ID", _next_id())
            item.set("CompositionName", "Items")
            item_attrs = ET.SubElement(item, "AttributeList")
            ET.SubElement(item_attrs, "Culture").text = culture
            ET.SubElement(item_attrs, "Text").text = tag.comment[:200]  # TIA Portal limit

    ET.indent(root, space="  ")
    return '<?xml version="1.0" encoding="utf-8"?>\n' + ET.tostring(root, encoding="unicode")


# -- XLSX Generator -----------------------------------------------------------

def _write_xlsx(tags: list[PlcTag], dest: Path) -> bool:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tag Table"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    alt_fill    = PatternFill("solid", fgColor="EBF5FB")

    headers = ["Tag Name", "Data Type", "Address", "Signal Type", "Original Name", "Description"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, tag in enumerate(tags, start=2):
        ws.append([tag.name, tag.data_type, tag.address or "—",
                   tag.signal_type, tag.original, tag.comment])
        if row_idx % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = alt_fill

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 40

    ws.freeze_panes = "A2"
    wb.save(str(dest))
    return True


# -- Main Function ------------------------------------------------------------

def run_export(
    project_path: Path,
    table_name: Optional[str] = None,
    output_dir: Optional[Path] = None,
    write_xlsx: bool = True,
    timestamped: bool = True,
    name_source: str = "iec",
) -> ExportResult:
    """Generate the TIA tag table XML (+ optional XLSX).

    ``timestamped=False`` writes a stable file name so automated callers
    (send_to_tia) overwrite one file per project instead of accumulating
    a new XML on every transfer.

    ``name_source``:
      * ``"iec"``  — HW03 IEC names, RD01 fallback with IEC prefixes
                     (review/engineering artifact; CLI default)
      * ``"rd01"`` — raw RD01 signal names only, HW03 ignored — these are
                     the names the assembled OB1 actually references, so
                     this is what send_to_tia must import
    """
    result = ExportResult()

    if not project_path.exists():
        result.warnings.append(f"Project folder not found: {project_path}")
        return result

    # Tag source
    if name_source == "rd01":
        tags = _parse_rd01_direct(project_path, use_original=True)
        source = "RD01_IO_List.md (program names)"
        if not tags:
            result.warnings.append(
                "No tag data found — RD01_IO_List.md is missing or empty."
            )
            return result
    else:
        tags = _parse_hw03(project_path)
        if tags:
            source = "HW03_IEC_Tags.md"
        else:
            tags = _parse_rd01_direct(project_path)
            source = "RD01_IO_List.md (direct)"
            if not tags:
                result.warnings.append(
                    "No tag data found. The IEC Tag Generator (HW03) or "
                    "RD01_IO_List.md is required first."
                )
                return result

    result.warnings.append(f"Source: {source}")

    # Auto-generate missing addresses
    addr_gen = _AddrGen()
    auto_count = 0
    for tag in tags:
        if not tag.address:
            tag.address, _ = addr_gen.next(tag.signal_type)
            auto_count += 1
    result.auto_addressed = auto_count
    if auto_count:
        result.warnings.append(
            f"{auto_count} tags had no address — sequential auto-addresses assigned. "
            "Verify against physical wiring before importing into TIA Portal."
        )

    result.tags = tags

    # File name
    suffix    = f"_{datetime.now().strftime('%Y%m%d_%H%M')}" if timestamped else ""
    safe_name = re.sub(r"[^\w\-]", "_", project_path.name)
    tbl_name  = table_name or f"TIA_Tags_{safe_name}"
    tia_ver   = _load_tia_version(project_path)
    culture   = _load_comment_culture(project_path)

    out_dir = output_dir or (project_path / "_output")
    out_dir.mkdir(parents=True, exist_ok=True)

    # XML
    xml_str  = _build_xml(tags, tbl_name, tia_ver, culture=culture)
    xml_dest = out_dir / f"TIA_TagTable_{safe_name}{suffix}.xml"
    xml_dest.write_text(xml_str, encoding="utf-8")
    result.xml_path = xml_dest

    # XLSX (optional)
    if write_xlsx:
        xlsx_dest = out_dir / f"TIA_TagTable_{safe_name}{suffix}.xlsx"
        ok = _write_xlsx(tags, xlsx_dest)
        if ok:
            result.xlsx_path = xlsx_dest
        else:
            result.warnings.append("openpyxl not installed — XLSX skipped.")

    return result


def format_export_summary(result: ExportResult) -> str:
    lines = ["TIA Portal Tag Export Summary", ""]
    if result.xml_path:
        lines.append(f"  XML file   : {result.xml_path.name}")
    if result.xlsx_path:
        lines.append(f"  XLSX file  : {result.xlsx_path.name}")
    lines.append(f"  Total tags : {len(result.tags)}")
    if result.auto_addressed:
        lines.append(f"  Auto-addr  : {result.auto_addressed} tags (verification required)")
    if result.tags:
        counts: dict[str, int] = {}
        for t in result.tags:
            counts[t.signal_type] = counts.get(t.signal_type, 0) + 1
        for k, v in sorted(counts.items()):
            lines.append(f"    {k:<5}: {v}")
    if result.warnings:
        lines.append("")
        for w in result.warnings:
            lines.append(f"  {w}")
    return "\n".join(lines)


# -- CLI ----------------------------------------------------------------------

def main():
    import argparse
    p = argparse.ArgumentParser(description="TIA Portal Tag Table XML Export")
    p.add_argument("--project", metavar="PROJECT_PATH", required=True)
    p.add_argument("--out",   metavar="FOLDER",       help="Output folder")
    p.add_argument("--table", metavar="TABLE_NAME",   help="TIA tag table name")
    p.add_argument("--no-xlsx", action="store_true", help="Skip XLSX output")
    args = p.parse_args()

    result = run_export(
        Path(args.project),
        table_name=args.table,
        output_dir=Path(args.out) if args.out else None,
        write_xlsx=not args.no_xlsx,
    )
    print(format_export_summary(result))
    if result.xml_path:
        print(f"\nXML : {result.xml_path}")
    if result.xlsx_path:
        print(f"XLSX: {result.xlsx_path}")


if __name__ == "__main__":
    main()

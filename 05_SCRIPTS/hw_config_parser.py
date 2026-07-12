#!/usr/bin/env python3
"""
hw_config_parser.py — IO Physical Address Mapping Manager (Phase 22)

Tasks:
  1. Read _input/hardware_config.xlsx (EPLAN-style)
  2. Validate tags against RD01_IO_List.md
  3. Generate metadata/HW02_IO_Adresleme.md (for AI brief)
  4. Create an empty xlsx template (for the engineer to fill in)

CLI:
  python hw_config_parser.py --template _input/hardware_config.xlsx
  python hw_config_parser.py --parse    _input/hardware_config.xlsx
                              --rd01     metadata/RD01_IO_List.md
                              --out      metadata/HW02_IO_Adresleme.md
"""

from __future__ import annotations

import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional


# -- Column headers (case-insensitive) ----------------------------------------
COL_TAG      = "tag"
COL_IO_TYPE  = "io_type"        # DI / DQ / AI / AO
COL_MOD_TYPE = "module_type"     # ET200SP DI 16x24VDC ...
COL_RACK     = "rack"
COL_SLOT     = "slot"
COL_CHANNEL  = "channel"
COL_ADDR_NEW = "new_address"     # %I0.0 / %IW100 ...
COL_ADDR_OLD = "old_address"     # E 1.0 / I 1.0 ...
COL_DESC     = "description"

# Turkish/English alias -> internal key mapping
# (Turkish aliases kept on purpose to match legacy/Turkish customer Excel headers.)
HEADER_ALIASES: dict[str, str] = {
    # Tag
    "tag": COL_TAG, "tag adı": COL_TAG, "io tag": COL_TAG, "sembol": COL_TAG,
    # IO type
    "io tipi": COL_IO_TYPE, "io_tipi": COL_IO_TYPE, "tip": COL_IO_TYPE, "type": COL_IO_TYPE,
    "io type": COL_IO_TYPE,
    # Module type
    "modül tipi": COL_MOD_TYPE, "modul tipi": COL_MOD_TYPE, "module type": COL_MOD_TYPE,
    "module": COL_MOD_TYPE, "modül": COL_MOD_TYPE,
    # Rack
    "rack": COL_RACK, "raf": COL_RACK,
    # Slot
    "slot": COL_SLOT,
    # Channel
    "kanal": COL_CHANNEL, "channel": COL_CHANNEL, "ch": COL_CHANNEL, "ch.": COL_CHANNEL,
    # New address
    "yeni adres": COL_ADDR_NEW, "yeni_adres": COL_ADDR_NEW, "adres": COL_ADDR_NEW,
    "new address": COL_ADDR_NEW, "address": COL_ADDR_NEW, "%adres": COL_ADDR_NEW,
    "plc adres": COL_ADDR_NEW, "plc adresi": COL_ADDR_NEW,
    # Old address
    "eski adres": COL_ADDR_OLD, "eski_adres": COL_ADDR_OLD, "old address": COL_ADDR_OLD,
    "s5 adres": COL_ADDR_OLD, "s7-300 adres": COL_ADDR_OLD,
    # Description
    "açıklama": COL_DESC, "aciklama": COL_DESC, "description": COL_DESC,
    "desc": COL_DESC, "comment": COL_DESC,
}


@dataclass
class IOEntry:
    tag: str
    io_type: str            # DI / DQ / AI / AO
    module_type: str        # ET200SP DI 16x24VDC ...
    rack: str               # "0" or "—"
    slot: str               # "3"
    channel: str            # "0"
    addr_new: str           # %I0.0
    addr_old: str           # E 1.0
    description: str = ""
    row_num: int = 0        # Excel row number (for debugging)


@dataclass
class ParseResult:
    entries: list[IOEntry] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    source_file: str = ""


# -- Excel template generator -------------------------------------------------

def generate_template_xlsx(output_path: Path, sample_rows: int = 10) -> bool:
    """Create an empty EPLAN-style hardware_config.xlsx template."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed: pip install openpyxl", file=sys.stderr)
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IO_Physical_Address"

    # Header row definitions
    headers = [
        ("Tag",           "IO tag name (must match RD01)",          18),
        ("IO Type",       "DI / DQ / AI / AO",                      9),
        ("Module Type",   "ET200SP DI 16x24VDC / DQ 16x24VDC ...", 26),
        ("Rack",          "Rack number (usually 0)",                  7),
        ("Slot",          "Slot number (TIA Portal)",                 7),
        ("Channel",       "Channel number (starts at 0)",             8),
        ("New Address",   "%I0.0 / %Q0.0 / %IW100",                 14),
        ("Old Address",   "S5/S7-300: E1.0 / A2.0 / EW4",           14),
        ("Description",   "Free text (optional)",                    28),
    ]

    # Header colors
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    hint_fill   = PatternFill("solid", fgColor="D6E4F7")
    hint_font   = Font(color="2F5496", italic=True, name="Calibri", size=9)
    border = Border(
        bottom=Side(style="thin"),
        right=Side(style="thin"),
    )

    # Column widths + header + sub-header (hint)
    for col_idx, (header, hint, width) in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

        # Header cell (row 1)
        cell1 = ws.cell(row=1, column=col_idx, value=header)
        cell1.font = header_font
        cell1.fill = header_fill
        cell1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell1.border = border

        # Hint cell (row 2)
        cell2 = ws.cell(row=2, column=col_idx, value=hint)
        cell2.font = hint_font
        cell2.fill = hint_fill
        cell2.alignment = Alignment(horizontal="left", vertical="center")
        cell2.border = border

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 16

    # Sample rows
    samples = [
        ("DI_EMSTOP_001",  "DI", "ET200SP DI 16x24VDC HF", "0", "3", "0",  "%I0.0",   "E 0.0", "Emergency stop button"),
        ("DI_DOOR_SW_001", "DI", "ET200SP DI 16x24VDC HF", "0", "3", "1",  "%I0.1",   "E 0.1", "Door switch"),
        ("DI_MOT1_FB",     "DI", "ET200SP DI 16x24VDC HF", "0", "3", "2",  "%I0.2",   "E 0.2", "Motor 1 feedback"),
        ("DI_PROX1",       "DI", "ET200SP DI 16x24VDC HF", "0", "3", "3",  "%I0.3",   "E 0.3", "Inductive sensor 1"),
        ("DQ_MOT1_KNT",    "DQ", "ET200SP DQ 16x24VDC HF", "0", "4", "0",  "%Q0.0",   "A 0.0", "Motor 1 contactor"),
        ("DQ_ALARM_LAMP",  "DQ", "ET200SP DQ 16x24VDC HF", "0", "4", "1",  "%Q0.1",   "A 0.1", "Alarm lamp"),
        ("AI_PRESS_001",   "AI", "ET200SP AI 4xU/I/RTD ST", "0", "5", "0", "%IW100",  "EW 4",  "Pressure transmitter 4-20mA"),
        ("AI_TEMP_001",    "AI", "ET200SP AI 4xU/I/RTD ST", "0", "5", "1", "%IW102",  "EW 6",  "PT100 temperature sensor"),
    ]

    data_fill_odd  = PatternFill("solid", fgColor="FFFFFF")
    data_fill_even = PatternFill("solid", fgColor="EBF5FB")
    data_font = Font(name="Calibri", size=10)

    for s_idx, sample in enumerate(samples[:sample_rows]):
        row_idx = s_idx + 3
        fill = data_fill_even if s_idx % 2 == 0 else data_fill_odd
        for col_idx, val in enumerate(sample, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Empty rows (to be filled in)
    for extra_row in range(3 + len(samples), 3 + len(samples) + max(0, sample_rows - len(samples)) + 50):
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=extra_row, column=col_idx).border = border

    # Info sheet
    info_ws = wb.create_sheet("Info")
    info_ws["A1"] = "AUTOMATION_FACTORY — IO Physical Address Configuration"
    info_ws["A1"].font = Font(bold=True, size=13)
    info_rows = [
        ("", ""),
        ("Purpose:", "This file maps RD01_IO_List.md to the TIA Portal IO configuration."),
        ("", "When generating AI code it uses real addresses like %I0.0 — not placeholders."),
        ("", ""),
        ("Rule 1:", "The Tag column must be IDENTICAL to the tag names in RD01_IO_List.md."),
        ("Rule 2:", "Take the New Address column from TIA Portal Hardware Config (%I0.0 format)."),
        ("Rule 3:", "The Old Address column is optional — for analysis of an S5/S7-300 project."),
        ("Rule 4:", "Channel numbers start at 0 (TIA Portal convention)."),
        ("", ""),
        ("IO Type:", "DI = Digital Input, DQ = Digital Output, AI = Analog Input, AO = Analog Output"),
        ("Address ex:", "DI/DQ -> %I0.0 or %Q0.0,  AI/AO -> %IW100 or %QW100"),
        ("Old Address:", "S5: E 0.0 (input) / A 0.0 (output),  S7-300: I 0.0 / Q 0.0"),
    ]
    for r_idx, (label, text) in enumerate(info_rows, start=2):
        if label:
            cell = info_ws.cell(row=r_idx, column=1, value=label)
            cell.font = Font(bold=True, size=10)
        info_ws.cell(row=r_idx, column=2, value=text).font = Font(size=10)
    info_ws.column_dimensions["A"].width = 14
    info_ws.column_dimensions["B"].width = 90

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return True


# -- Excel parser -------------------------------------------------------------

def _normalize_header(raw: str) -> str:
    return str(raw).strip().lower().replace("\n", " ").replace("_", " ")


def parse_hw_config(xlsx_path: Path) -> ParseResult:
    """Read hardware_config.xlsx -> return ParseResult."""
    result = ParseResult(source_file=str(xlsx_path))

    try:
        import openpyxl
    except ImportError:
        result.errors.append("openpyxl not installed: pip install openpyxl")
        return result

    if not xlsx_path.exists():
        result.errors.append(f"File not found: {xlsx_path}")
        return result

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception as e:
        result.errors.append(f"Excel could not be opened: {e}")
        return result

    # Use first sheet (or the IO_Physical_Address-named sheet; legacy substrings kept)
    ws = None
    for sheet_name in wb.sheetnames:
        low = sheet_name.lower()
        if ("address" in low or "physical" in low or "io" in low
                or "adres" in low or "fiziksel" in low):
            ws = wb[sheet_name]
            break
    if ws is None:
        ws = wb.active

    rows_iter = list(ws.iter_rows(values_only=True))
    if not rows_iter:
        result.errors.append("Excel is empty")
        return result

    # Find header row (row 1 or 2 — might be hint row)
    header_row_idx = None
    col_map: dict[str, int] = {}  # internal_key -> column_index (0-based)

    for candidate_idx in (0, 1):
        if candidate_idx >= len(rows_iter):
            break
        row = rows_iter[candidate_idx]
        tentative: dict[str, int] = {}
        for col_idx, cell_val in enumerate(row):
            if cell_val is None:
                continue
            norm = _normalize_header(str(cell_val))
            if norm in HEADER_ALIASES:
                tentative[HEADER_ALIASES[norm]] = col_idx
        if COL_TAG in tentative:
            header_row_idx = candidate_idx
            col_map = tentative
            break

    if header_row_idx is None:
        result.errors.append(
            "Header row not found. 'Tag' column is required. "
            "Found headers: " + str([r for r in rows_iter[0] if r])
        )
        return result

    if COL_ADDR_NEW not in col_map:
        result.warnings.append(
            "'New Address' column not found — physical address mapping missing."
        )

    # Data rows (skip header + possible hint row)
    data_start = header_row_idx + 1
    if header_row_idx == 0 and len(rows_iter) > 1:
        # Is row 2 a hint? (colored but not data)
        sample_row = rows_iter[1]
        tag_col = col_map.get(COL_TAG, 0)
        if tag_col < len(sample_row):
            tag_val = str(sample_row[tag_col] or "").strip()
            # If tag is "DI", "DQ" or a short description, treat as hint row
            if len(tag_val) < 5 or tag_val.lower() in ("tag", "tag adı", "io tag", "sembol", "—", "-"):
                data_start = 2

    for row_offset, row in enumerate(rows_iter[data_start:], start=data_start + 1):
        def get(col_key: str, default: str = "") -> str:
            idx = col_map.get(col_key)
            if idx is None or idx >= len(row):
                return default
            val = row[idx]
            return str(val).strip() if val is not None else default

        tag = get(COL_TAG)
        if not tag or tag.lower() in ("tag", "tag adı", "—", "-", "none", ""):
            continue  # skip empty row

        entry = IOEntry(
            tag=tag,
            io_type=get(COL_IO_TYPE, "?").upper(),
            module_type=get(COL_MOD_TYPE, "—"),
            rack=get(COL_RACK, "0"),
            slot=get(COL_SLOT, "?"),
            channel=get(COL_CHANNEL, "?"),
            addr_new=get(COL_ADDR_NEW, "—"),
            addr_old=get(COL_ADDR_OLD, "—"),
            description=get(COL_DESC, ""),
            row_num=row_offset,
        )
        result.entries.append(entry)

    if not result.entries:
        result.warnings.append("No data rows found — delete the template example rows and enter the real IO list.")

    return result


# -- RD01 tag extraction ------------------------------------------------------

def extract_rd01_tags(rd01_path: Path) -> list[str]:
    """Extract the tag list from RD01_IO_List.md."""
    if not rd01_path.exists():
        return []
    tags = []
    content = rd01_path.read_text(encoding="utf-8")
    # Find the first column separated by | in table rows
    # Format: | TAG_NAME | DI | ... |
    tag_pattern = re.compile(
        r"^\s*\|\s*([A-Z][A-Z0-9_]{2,})\s*\|",
        re.MULTILINE,
    )
    for m in tag_pattern.finditer(content):
        tag = m.group(1).strip()
        # Only add if it is not a header row
        if tag.upper() not in ("TAG", "SEMBOL", "NO", "ID", "NAME", "AD"):
            tags.append(tag)
    return tags


# -- Validation ---------------------------------------------------------------

def validate_against_rd01(
    hw_entries: list[IOEntry],
    rd01_path: Path,
) -> dict:
    """Compare hardware_config tags with RD01 tags."""
    rd01_tags = set(extract_rd01_tags(rd01_path))
    hw_tags = {e.tag for e in hw_entries}

    only_in_hw  = hw_tags - rd01_tags
    only_in_rd01 = rd01_tags - hw_tags
    common = hw_tags & rd01_tags

    return {
        "rd01_count": len(rd01_tags),
        "hw_count":   len(hw_tags),
        "matched":    len(common),
        "only_in_hw": sorted(only_in_hw),
        "only_in_rd01": sorted(only_in_rd01),
        "ok": len(only_in_hw) == 0 and len(only_in_rd01) == 0,
    }


# -- HW02 MD generator --------------------------------------------------------

def generate_hw02_md(
    entries: list[IOEntry],
    output_path: Path,
    source_xlsx: Path,
    validation: Optional[dict] = None,
) -> bool:
    """Generate metadata/HW02_IO_Adresleme.md."""
    from datetime import datetime

    lines = []
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")

    lines.append("# HW02 — IO Physical Address Mapping")
    lines.append("")
    lines.append("```yaml")
    lines.append("document_id: HW02")
    lines.append("title: IO Physical Address Mapping")
    lines.append(f"source: {source_xlsx.name}")
    lines.append(f"updated: {ts}")
    lines.append(f"entry_count: {len(entries)}")
    if validation:
        v_status = "OK" if validation["ok"] else "MISMATCH"
        lines.append(f"rd01_match: {v_status}  # rd01={validation['rd01_count']} hw={validation['hw_count']} matched={validation['matched']}")
    lines.append("```")
    lines.append("")

    # Validation warnings
    if validation and not validation["ok"]:
        lines.append("## Validation Warnings")
        if validation["only_in_hw"]:
            lines.append(f"\n**In hardware_config but NOT in RD01** ({len(validation['only_in_hw'])} items):")
            for t in validation["only_in_hw"]:
                lines.append(f"  - `{t}`")
        if validation["only_in_rd01"]:
            lines.append(f"\n**In RD01 but NOT in hardware_config** ({len(validation['only_in_rd01'])} items):")
            for t in validation["only_in_rd01"]:
                lines.append(f"  - `{t}`")
        lines.append("")

    # DI table
    di_entries = [e for e in entries if e.io_type in ("DI", "BOOL")]
    if di_entries:
        lines.append("## Digital Inputs (DI)")
        lines.append("")
        lines.append("| Tag | New Address | Rack/Slot/Ch | Module | Old Address | Description |")
        lines.append("|-----|-------------|--------------|--------|-------------|-------------|")
        for e in di_entries:
            lines.append(
                f"| `{e.tag}` | `{e.addr_new}` | {e.rack}/{e.slot}/{e.channel} "
                f"| {e.module_type} | `{e.addr_old}` | {e.description} |"
            )
        lines.append("")

    # DQ table
    dq_entries = [e for e in entries if e.io_type in ("DQ", "DO")]
    if dq_entries:
        lines.append("## Digital Outputs (DQ)")
        lines.append("")
        lines.append("| Tag | New Address | Rack/Slot/Ch | Module | Old Address | Description |")
        lines.append("|-----|-------------|--------------|--------|-------------|-------------|")
        for e in dq_entries:
            lines.append(
                f"| `{e.tag}` | `{e.addr_new}` | {e.rack}/{e.slot}/{e.channel} "
                f"| {e.module_type} | `{e.addr_old}` | {e.description} |"
            )
        lines.append("")

    # AI table
    ai_entries = [e for e in entries if e.io_type in ("AI", "ANALOG_IN")]
    if ai_entries:
        lines.append("## Analog Inputs (AI)")
        lines.append("")
        lines.append("| Tag | New Address | Rack/Slot/Ch | Module | Old Address | Description |")
        lines.append("|-----|-------------|--------------|--------|-------------|-------------|")
        for e in ai_entries:
            lines.append(
                f"| `{e.tag}` | `{e.addr_new}` | {e.rack}/{e.slot}/{e.channel} "
                f"| {e.module_type} | `{e.addr_old}` | {e.description} |"
            )
        lines.append("")

    # AO table
    ao_entries = [e for e in entries if e.io_type in ("AO", "ANALOG_OUT")]
    if ao_entries:
        lines.append("## Analog Outputs (AO)")
        lines.append("")
        lines.append("| Tag | New Address | Rack/Slot/Ch | Module | Old Address | Description |")
        lines.append("|-----|-------------|--------------|--------|-------------|-------------|")
        for e in ao_entries:
            lines.append(
                f"| `{e.tag}` | `{e.addr_new}` | {e.rack}/{e.slot}/{e.channel} "
                f"| {e.module_type} | `{e.addr_old}` | {e.description} |"
            )
        lines.append("")

    # Unclassified
    other = [e for e in entries if e.io_type not in ("DI", "DQ", "AI", "AO", "DO", "BOOL", "ANALOG_IN", "ANALOG_OUT")]
    if other:
        lines.append("## Other / Unclassified")
        lines.append("")
        lines.append("| Tag | IO Type | New Address | Old Address | Description |")
        lines.append("|-----|---------|-------------|-------------|-------------|")
        for e in other:
            lines.append(
                f"| `{e.tag}` | {e.io_type} | `{e.addr_new}` | `{e.addr_old}` | {e.description} |"
            )
        lines.append("")

    # Summary statistics
    lines.append("## Summary")
    lines.append("")
    lines.append(f"- **Total IO:** {len(entries)}")
    lines.append(f"- Digital Inputs (DI): {len(di_entries)}")
    lines.append(f"- Digital Outputs (DQ): {len(dq_entries)}")
    lines.append(f"- Analog Inputs (AI): {len(ai_entries)}")
    lines.append(f"- Analog Outputs (AO): {len(ao_entries)}")
    if other:
        lines.append(f"- Other: {len(other)}")
    if validation:
        lines.append(f"- **RD01 Match:** {'Perfect match' if validation['ok'] else 'Mismatch present'} "
                     f"({validation['matched']}/{max(validation['rd01_count'], validation['hw_count'])})")
    lines.append("")
    lines.append("---")
    lines.append(f"*Auto-generated: AUTOMATION_FACTORY hw_config_parser.py — {ts}*")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines), encoding="utf-8")
    return True


# -- CLI interface ------------------------------------------------------------

def main():
    import argparse
    parser = argparse.ArgumentParser(description="IO Hardware Config Parser")
    parser.add_argument("--template", metavar="PATH",
                        help="Create empty xlsx template (e.g.: _input/hardware_config.xlsx)")
    parser.add_argument("--parse", metavar="PATH",
                        help="Process a hardware_config.xlsx file")
    parser.add_argument("--rd01", metavar="PATH",
                        help="RD01_IO_List.md path (for validation)")
    parser.add_argument("--out", metavar="PATH",
                        help="HW02 output MD path (default: metadata/HW02_IO_Adresleme.md)")
    args = parser.parse_args()

    if args.template:
        out = Path(args.template)
        print(f"Creating template: {out}")
        ok = generate_template_xlsx(out)
        if ok:
            print(f"Template created: {out}")
        else:
            print("Template could not be created (is openpyxl installed?)")
        return

    if args.parse:
        xlsx_path = Path(args.parse)
        result = parse_hw_config(xlsx_path)

        if result.errors:
            for e in result.errors:
                print(f"ERROR: {e}")
            return

        print(f"Read {len(result.entries)} IO entries: {xlsx_path}")

        if result.warnings:
            for w in result.warnings:
                print(f"WARN: {w}")

        validation = None
        if args.rd01:
            rd01_path = Path(args.rd01)
            validation = validate_against_rd01(result.entries, rd01_path)
            if validation["ok"]:
                print(f"RD01 match: {validation['matched']} tags — perfect match")
            else:
                print(f"RD01 mismatch:")
                if validation["only_in_hw"]:
                    print(f"   In HW, NOT in RD01: {validation['only_in_hw']}")
                if validation["only_in_rd01"]:
                    print(f"   In RD01, NOT in HW: {validation['only_in_rd01']}")

        out_path = Path(args.out) if args.out else Path("metadata/HW02_IO_Adresleme.md")
        generate_hw02_md(result.entries, out_path, xlsx_path, validation)
        print(f"HW02 generated: {out_path}")
        return

    parser.print_help()


if __name__ == "__main__":
    main()

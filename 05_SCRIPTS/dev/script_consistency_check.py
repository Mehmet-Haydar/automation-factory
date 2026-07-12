#!/usr/bin/env python3
"""
script_consistency_check.py
=============================
Checks that files in the factory or a project conform to the naming standard.

USAGE:
    # Audit the factory (default)
    python script_consistency_check.py

    # Audit a project folder (for PLC tags)
    python script_consistency_check.py --project ~/projects/MyProject \\
                                        --check-naming --check-addresses

CHECKED:
    1. Factory .md files -> name format [SCOPE]_[DOMAIN]_[SUB].md
    2. Project IO list (CSV/XLSX) -> tag format TYPE_LOC_NUM_FUNC

OUTPUT:
    List of files/tags that do not conform to the standard.
"""

import argparse
import re
import sys
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# This script lives in 05_SCRIPTS/dev/, so the repo root is three levels up.
FACTORY_ROOT = Path(__file__).resolve().parent.parent.parent

# Factory file-name regex
FACTORY_FILENAME_RE = re.compile(
    r"^(GLOBAL|RETROFIT|GREENFIELD|DOMAIN|PROMPT|SCRIPT|KB|FACTORY|PROJECT|README|CHANGELOG|MDSCHEMA)"
    r"(_[A-Z][A-Z0-9_]*)*"
    r"\.(md|scl|sh|py|json)$"
)

# PLC tag regex (for project audit)
PLC_TAG_RE = re.compile(
    r"^(MOT|VLV|SNS|PRX|ENC|TMP|LVL|PRS|FLW|PB|LMP|ESD|LS)"
    r"_[A-Z]{2,4}\d{2}"          # LOC: 2-4 letters + 2 digits (CV01, AX01) or
    r"|^(MOT|VLV|SNS|PRX|ENC|TMP|LVL|PRS|FLW|PB|LMP|ESD|LS)"
    r"_(PNL|HMI)"                # special locations
    r"_\d{3}"                    # NUM: 3 digits
    r"_[A-Z][A-Z0-9_]*$"          # FUNC
)

# Allowed domain-specific filenames (near the factory root)
ALLOWED_TOP_LEVEL = {
    "FACTORY_MAESTRO.md", "README.md", "CHANGELOG.md",
    "PROJECT_VISION.md", "SKELETON_BLUEPRINT.md", "PROGRESS_TRACKER.md",
    ".cursorrules", ".gitignore",
}

# Meta files (may also exist at the folder level)
META_FILES = {
    "METADATA_INPUT_GUIDE.md",
    "PROJECT_MAESTRO_TEMPLATE.md",
    "PROJECT_STATE_TEMPLATE.json",
}


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--project", type=Path,
                   help="Project folder (if given, audits the project instead of the factory)")
    p.add_argument("--check-naming", action="store_true",
                   help="Check PLC tag naming")
    p.add_argument("--check-addresses", action="store_true",
                   help="Check PLC address conflicts (TODO)")
    p.add_argument("--io-file", type=Path,
                   help="IO list (CSV/XLSX). If omitted, searched under 03_PLC.")
    return p.parse_args()


# -----------------------------------------------------------------------------
# Factory audit
# -----------------------------------------------------------------------------

def check_factory() -> int:
    """Audit factory file names."""
    issues = []
    skipped_dirs = {".git", ".cursor", "node_modules", "__pycache__",
                    ".venv", "venv", "_archive", ".pytest_cache",
                    ".mypy_cache", ".ruff_cache", "site-packages"}

    for path in FACTORY_ROOT.rglob("*"):
        if not path.is_file():
            continue
        # Skip ignored
        if any(part in skipped_dirs for part in path.parts):
            continue
        # Skip 07_PROJECT_TEMPLATE — copied to the user, flexible
        if "07_PROJECT_TEMPLATE" in path.parts:
            continue

        rel = path.relative_to(FACTORY_ROOT)
        name = path.name

        # Skip top-level exceptions
        if rel.parts == (name,) and name in ALLOWED_TOP_LEVEL:
            continue

        # Meta files (may exist in any folder)
        if name in META_FILES:
            continue

        # .mdc files under .cursor/rules/ use a different format
        if path.suffix == ".mdc":
            continue

        # Only check .md, .scl, .sh, .py
        if path.suffix not in {".md", ".scl", ".sh", ".py"}:
            continue

        # Script files follow a different rule — script_<name>.py
        if path.suffix in {".py", ".sh"}:
            if not name.startswith("script_"):
                issues.append((str(rel), "Script file must start with 'script_'"))
            continue

        if not FACTORY_FILENAME_RE.match(name):
            issues.append((str(rel), "Name format: SCOPE_DOMAIN_SUB.ext"))

    print(f"\nFactory Naming Consistency Check")
    print(f"   Root: {FACTORY_ROOT}")
    print(f"   Issues found: {len(issues)}\n")

    if not issues:
        print("All filenames conform to naming standard.")
        return 0

    print(f"{'File':<70} Issue")
    print("-" * 100)
    for path, reason in issues:
        print(f"{path:<70} {reason}")

    return 1


# -----------------------------------------------------------------------------
# Project audit (PLC tags)
# -----------------------------------------------------------------------------

def find_io_file(project_root: Path) -> Path | None:
    """Find the IO list inside a project."""
    candidates = [
        project_root / "03_PLC" / "IO_RAW_FROM_EPLAN.xlsx",
        project_root / "03_PLC" / "IO_LIST.xlsx",
        project_root / "03_PLC" / "IO_LIST.csv",
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


def check_plc_tags(io_file: Path) -> int:
    """Audit PLC tags."""
    print(f"\nPLC Tag Consistency Check")
    print(f"   File: {io_file}")

    if io_file.suffix == ".csv":
        import csv
        with io_file.open(encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
    else:
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("openpyxl required. pip install openpyxl")
            return 1
        wb = load_workbook(io_file, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = list(next(rows_iter, []))
        rows = [dict(zip(headers, r)) for r in rows_iter]

    if not rows:
        print("IO list is empty.")
        return 0

    # Find the tag column name
    sample = rows[0]
    tag_col = None
    for candidate in ("Tag", "tag", "TAG", "Name", "Symbol"):
        if candidate in sample:
            tag_col = candidate
            break
    if not tag_col:
        print(f"Tag column not found. Columns: {list(sample.keys())}")
        return 1

    bad = []
    for i, row in enumerate(rows, start=2):  # header was row 1
        tag = row.get(tag_col)
        if not tag:
            continue
        if not PLC_TAG_RE.match(str(tag).strip()):
            bad.append((i, tag))

    print(f"   Total tags: {len(rows)}")
    print(f"   Bad tags  : {len(bad)}\n")

    if not bad:
        print("All tags conform to naming standard.")
        return 0

    print(f"{'Row':<6} {'Tag':<40} Issue")
    print("-" * 80)
    for row_num, tag in bad[:50]:  # first 50
        print(f"{row_num:<6} {tag!s:<40} Name format: TYPE_LOC_NUM_FUNC")
    if len(bad) > 50:
        print(f"... and {len(bad) - 50} more")

    return 1


def main() -> int:
    args = parse_args()

    if args.project:
        if not args.project.exists():
            print(f"Project not found: {args.project}")
            return 1
        if args.check_naming:
            io_file = args.io_file or find_io_file(args.project)
            if io_file is None:
                print("IO file not found. Provide the path with --io-file.")
                return 1
            return check_plc_tags(io_file)
        else:
            print("--check-naming or --check-addresses required.")
            return 1

    # Default: factory audit
    return check_factory()


if __name__ == "__main__":
    sys.exit(main())

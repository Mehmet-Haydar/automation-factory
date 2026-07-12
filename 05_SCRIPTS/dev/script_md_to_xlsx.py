#!/usr/bin/env python3
"""
script_md_to_xlsx.py — Per-project MD template -> XLSX conversion

Dependencies: openpyxl, pyyaml
Output: XLSX file (per-project metadata_template)
Sandbox: write (creates a new file)

Usage:
    python 05_SCRIPTS/dev/script_md_to_xlsx.py \\
        --input  07_PROJECT_TEMPLATE/metadata_template/RD01_IO_List.md \\
        --output 07_PROJECT_TEMPLATE/metadata_template/RD01_IO_List.xlsx

    # Batch-convert every RD:
    python 05_SCRIPTS/dev/script_md_to_xlsx.py --batch
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

try:
    import yaml  # type: ignore
    from openpyxl import Workbook  # type: ignore
    from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore
except ImportError as e:
    print(f"[ERR] Missing dependency: {e}. Run `pip install openpyxl pyyaml`.", file=sys.stderr)
    sys.exit(2)


def parse_md_template(md_path: Path) -> dict:
    """
    Parse an MD template:
    - Frontmatter
    - Table headers (column names)
    - Table rows (example data)
    """
    text = md_path.read_text(encoding="utf-8")

    # Frontmatter
    fm_match = re.match(r"^---\n(.*?)\n---\n", text, re.DOTALL)
    frontmatter = yaml.safe_load(fm_match.group(1)) if fm_match else {}

    # Tables: after a | col1 | col2 | ... | line
    tables = []
    current_table = None
    in_table = False
    skip_separator = False

    for line in text.splitlines():
        if line.startswith("|") and not skip_separator:
            cells = [c.strip() for c in line.strip("|").split("|")]
            if not in_table:
                current_table = {"header": cells, "rows": []}
                in_table = True
                skip_separator = True
            else:
                current_table["rows"].append(cells)
        elif line.startswith("|---") or line.startswith("|-"):
            skip_separator = False
            continue
        else:
            if in_table and current_table:
                tables.append(current_table)
                current_table = None
                in_table = False
                skip_separator = False

    if in_table and current_table:
        tables.append(current_table)

    return {"frontmatter": frontmatter, "tables": tables}


def write_xlsx(parsed: dict, output_path: Path) -> None:
    """Parsed MD -> XLSX file."""
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Each table goes on its own sheet
    for idx, table in enumerate(parsed["tables"], start=1):
        sheet_name = f"Sheet{idx}"
        # Pick a meaningful name
        if table["header"]:
            first_col = table["header"][0].lower()
            if "tag" in first_col or "signal" in first_col:
                sheet_name = "Signals"
            elif "varname" in first_col:
                sheet_name = "Variables"
            elif "stepid" in first_col:
                sheet_name = "Steps"
            elif "modeid" in first_col:
                sheet_name = "Modes"
            elif "functionid" in first_col:
                sheet_name = "Functions"
            elif "axisid" in first_col:
                sheet_name = "Axes"
            elif "timerid" in first_col:
                sheet_name = "Timers"
            elif "alarmid" in first_col:
                sheet_name = "Alarms"
            elif "commid" in first_col:
                sheet_name = "Connections"
            elif "blockname" in first_col:
                sheet_name = "BlockList" if "BlockType" in table["header"] else "ParamList"
            elif "screenid" in first_col:
                sheet_name = "ScreenList"
            elif "hmi_tagid" in first_col:
                sheet_name = "TagList"
            elif "usecaseid" in first_col:
                sheet_name = "UseCases"
            elif "annotationid" in first_col:
                sheet_name = "Annotations"
            elif "findingid" in first_col:
                sheet_name = "Findings"
            elif "category" in first_col:
                sheet_name = "Summary"

        ws = wb.create_sheet(sheet_name[:31])  # Excel limit: 31 chars

        # Header row (formatted)
        for col_idx, header in enumerate(table["header"], start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        for row_idx, row in enumerate(table["rows"], start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Column width auto
        for col_idx, header in enumerate(table["header"], start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(len(header) + 2, 15)

    # Frontmatter sheet (meta)
    fm = parsed["frontmatter"]
    if fm:
        meta_ws = wb.create_sheet("_Meta", 0)
        meta_ws["A1"] = "Field"
        meta_ws["B1"] = "Value"
        meta_ws["A1"].font = Font(bold=True)
        meta_ws["B1"].font = Font(bold=True)
        for row_idx, (key, val) in enumerate(fm.items(), start=2):
            meta_ws.cell(row=row_idx, column=1, value=str(key))
            meta_ws.cell(row=row_idx, column=2, value=str(val))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def find_all_templates(root: Path) -> list[Path]:
    """Find every MD file inside 07_PROJECT_TEMPLATE/metadata_template/."""
    template_dir = root / "07_PROJECT_TEMPLATE" / "metadata_template"
    return sorted(template_dir.glob("RD*.md"))


def main() -> int:
    parser = argparse.ArgumentParser(description="MD template -> XLSX conversion")
    parser.add_argument("--input", type=Path, help="Single MD file input")
    parser.add_argument("--output", type=Path, help="XLSX output path")
    parser.add_argument("--batch", action="store_true", help="Convert every metadata_template MD")
    parser.add_argument("--dry-run", action="store_true", help="Parse only, do not write")
    args = parser.parse_args()

    # This script lives in 05_SCRIPTS/dev/, so the repo root is three levels up.
    repo_root = Path(__file__).resolve().parent.parent.parent

    if args.batch:
        templates = find_all_templates(repo_root)
        print(f"[INFO] Found {len(templates)} MD templates")
        for md_path in templates:
            xlsx_path = md_path.with_suffix(".xlsx")
            print(f"  {md_path.name} -> {xlsx_path.name}")
            if not args.dry_run:
                parsed = parse_md_template(md_path)
                write_xlsx(parsed, xlsx_path)
        return 0

    if not args.input or not args.output:
        parser.error("--input and --output required (or use --batch)")

    parsed = parse_md_template(args.input)
    print(f"[INFO] {len(parsed['tables'])} tables found, frontmatter: {len(parsed['frontmatter'])} fields")

    if not args.dry_run:
        write_xlsx(parsed, args.output)
        print(f"[OK] Written: {args.output}")
    else:
        print("[DRY-RUN] Write skipped")

    return 0


if __name__ == "__main__":
    sys.exit(main())

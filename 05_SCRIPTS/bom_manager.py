#!/usr/bin/env python3
"""
bom_manager.py — Hardware BOM (Bill of Materials) Manager (Phase 23)

Tasks:
  1. Generate a BOM Excel from the selected_devices list in PROJECT_STATE.json
  2. BOM Excel -> metadata/HW01_BOM.md (for the AI brief)
  3. Read each device's MD from the library -> embedded into the brief

CLI:
  python bom_manager.py --generate  PROJECT_PATH
  python bom_manager.py --hw01      PROJECT_PATH
"""

from __future__ import annotations

import json
import logging
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional


FACTORY_ROOT = Path(__file__).resolve().parent.parent
HW_LIB_ROOT  = FACTORY_ROOT / "09_HARDWARE_LIBRARY"

# O-1 fix: module-local logger so file-read failures in catalog / BOM lookup
# are visible in stderr/log rather than silently skipped.
_logger = logging.getLogger("bom_manager")


def _warn(msg: str) -> None:
    """Emit a recoverable warning to stderr via logging."""
    _logger.warning("[parse] %s", msg)


@dataclass
class BOMEntry:
    device_id: str
    vendor: str
    model: str
    category: str
    quantity: int
    notes: str = ""
    md_path: Optional[Path] = None
    md_content: str = ""


# -- Library scan -------------------------------------------------------------

def scan_library(lib_root: Path) -> dict[str, dict]:
    """Scan the library MDs -> returns {device_id: {vendor, model, category, path}}."""
    catalog: dict[str, dict] = {}
    if not lib_root.exists():
        return catalog
    for md_path in sorted(lib_root.rglob("*.md")):
        if md_path.name.startswith("_"):
            continue
        parts = md_path.relative_to(lib_root).parts
        category = parts[0] if len(parts) > 1 else "accessories"
        device_id = vendor = model = ""
        try:
            content = md_path.read_text(encoding="utf-8")
            for line in content.splitlines():
                line = line.strip()
                if line.startswith("device_id:"):
                    device_id = line.split(":", 1)[1].strip().strip('"')
                elif line.startswith("vendor:"):
                    vendor = line.split(":", 1)[1].strip().strip('"')
                elif line.startswith("model:"):
                    model = line.split(":", 1)[1].strip().strip('"')
                if device_id and vendor and model:
                    break
        except Exception as exc:
            # O-1 fix: file-read failure is logged; catalog entry is skipped.
            _warn(f"scan_library: cannot read {md_path}: {exc}")
        if device_id:
            catalog[device_id] = {
                "vendor": vendor,
                "model": model,
                "category": category,
                "path": md_path,
            }
    return catalog


def find_device_md(device_id: str, project_path: Path) -> Optional[Path]:
    """Find the device MD, first in the project-local folder, then in the factory library."""
    # 1. Project-local _hardware/
    if project_path:
        for md in (project_path / "_hardware").rglob("*.md"):
            if md.name.startswith("_"):
                continue
            try:
                content = md.read_text(encoding="utf-8")
                for line in content.splitlines():
                    if line.strip().startswith("device_id:"):
                        did = line.split(":", 1)[1].strip().strip('"')
                        if did == device_id:
                            return md
            except Exception as exc:
                # O-1 fix: file-read failure is logged; search continues.
                _warn(f"find_device_md: cannot read {md}: {exc}")
    # 2. Factory library
    for md in HW_LIB_ROOT.rglob("*.md"):
        if md.name.startswith("_"):
            continue
        try:
            content = md.read_text(encoding="utf-8")
            for line in content.splitlines():
                if line.strip().startswith("device_id:"):
                    did = line.split(":", 1)[1].strip().strip('"')
                    if did == device_id:
                        return md
        except Exception as exc:
            # O-1 fix: file-read failure is logged; search continues.
            _warn(f"find_device_md: cannot read {md}: {exc}")
    return None


# -- BOM Excel generator ------------------------------------------------------

def generate_bom_xlsx(
    project_path: Path,
    selected_devices: list[dict],
    catalog: Optional[dict[str, dict]] = None,
) -> bool:
    """Generate _input/hardware_BOM.xlsx."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed: pip install openpyxl", file=sys.stderr)
        return False

    if catalog is None:
        catalog = scan_library(HW_LIB_ROOT)
        # Add project-local too
        proj_hw = project_path / "_hardware"
        if proj_hw.exists():
            catalog.update(scan_library(proj_hw))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hardware_BOM"

    # Header style
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    hint_fill = PatternFill("solid", fgColor="D6E4F7")
    hint_font = Font(color="2F5496", italic=True, name="Calibri", size=9)
    border = Border(
        bottom=Side(style="thin"), right=Side(style="thin")
    )
    data_font = Font(name="Calibri", size=10)
    even_fill = PatternFill("solid", fgColor="EBF5FB")
    odd_fill  = PatternFill("solid", fgColor="FFFFFF")

    headers = [
        ("Device ID",    "Library identifier",            22),
        ("Vendor",       "Manufacturer name",             22),
        ("Model",        "Full model name",               30),
        ("Category",     "drives / io_modules / ...",     16),
        ("Quantity",     "Quantity in this project",       7),
        ("Notes",        "Project-specific notes (optional)", 32),
        ("Source",       "factory / project_local",       16),
    ]

    for col_idx, (hdr, hint, width) in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        c1 = ws.cell(row=1, column=col_idx, value=hdr)
        c1.font, c1.fill, c1.alignment, c1.border = (
            hdr_font, hdr_fill,
            Alignment(horizontal="center", vertical="center"), border
        )
        c2 = ws.cell(row=2, column=col_idx, value=hint)
        c2.font, c2.fill, c2.border = hint_font, hint_fill, border

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 14

    for row_offset, dev in enumerate(selected_devices):
        device_id = dev.get("device_id", "")
        quantity  = dev.get("quantity", 1)
        notes     = dev.get("notes", "")
        info = catalog.get(device_id, {})

        # Determine source
        proj_hw_path = project_path / "_hardware"
        src = "factory"
        if proj_hw_path.exists():
            for md in proj_hw_path.rglob("*.md"):
                if not md.name.startswith("_"):
                    try:
                        for line in md.read_text(encoding="utf-8").splitlines():
                            if line.strip().startswith("device_id:") and \
                               line.split(":", 1)[1].strip().strip('"') == device_id:
                                src = "project_local"
                    except Exception as exc:
                        # O-1 fix: file-read failure is logged; src defaults to factory.
                        _warn(f"generate_bom_xlsx: cannot read {md}: {exc}")

        row_idx = row_offset + 3
        fill = even_fill if row_offset % 2 == 0 else odd_fill
        row_data = [
            device_id,
            info.get("vendor", "—"),
            info.get("model", "—"),
            info.get("category", "—"),
            quantity,
            notes,
            src,
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font, cell.fill, cell.border = data_font, fill, border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # AUDIT-005: cross-check Safety IO vs Safety PLC presence.
    # Heuristic: any device whose category or model contains 'safety' or 'fail_safe'
    # is considered a Safety IO module; any device whose category or model contains
    # 'safety_cpu' or 'f-cpu' is considered a Safety PLC.
    _safety_io_ids: list[str] = []
    _has_safety_plc = False
    for dev in selected_devices:
        did = dev.get("device_id", "")
        info = catalog.get(did, {})
        cat   = (info.get("category") or "").lower()
        model = (info.get("model") or "").lower()
        combined = cat + " " + model
        if "safety_cpu" in combined or "f-cpu" in combined or "fcpu" in combined:
            _has_safety_plc = True
        elif "safety" in combined or "fail_safe" in combined or "failsafe" in combined:
            _safety_io_ids.append(did)
    if _safety_io_ids and not _has_safety_plc:
        warn_row = ws.max_row + 2
        warn_cell = ws.cell(
            row=warn_row, column=1,
            value=(
                "⚠ AUDIT-005 WARNING: Safety IO module(s) detected "
                f"({', '.join(_safety_io_ids)}) but no Safety PLC (F-CPU) found "
                "in this BOM. Verify hardware selection before ordering. (bom_manager)"
            )
        )
        from openpyxl.styles import Font as _Font, PatternFill as _Fill
        warn_cell.font = _Font(color="9C0006", bold=True, name="Calibri", size=10)
        warn_cell.fill = _Fill("solid", fgColor="FFC7CE")
        ws.merge_cells(
            start_row=warn_row, start_column=1,
            end_row=warn_row, end_column=len(headers)
        )

    out_path = project_path / "_input" / "hardware_BOM.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return True


# -- HW01_BOM.md generator ----------------------------------------------------

def generate_hw01_md(
    project_path: Path,
    selected_devices: list[dict],
    catalog: Optional[dict[str, dict]] = None,
    include_scl_snippets: bool = True,
) -> Path:
    """Generate metadata/HW01_BOM.md. Contains full device info for the AI brief."""
    if catalog is None:
        catalog = scan_library(HW_LIB_ROOT)
        proj_hw = project_path / "_hardware"
        if proj_hw.exists():
            catalog.update(scan_library(proj_hw))

    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    lines = []

    lines.append("# HW01 — Hardware BOM (Bill of Materials)")
    lines.append("")
    lines.append("```yaml")
    lines.append("document_id: HW01")
    lines.append("title: Project Hardware BOM")
    lines.append(f"updated: {ts}")
    lines.append(f"device_count: {len(selected_devices)}")
    lines.append("```")
    lines.append("")

    if not selected_devices:
        lines.append("_(No devices selected for this project yet. Add them via GUI -> Hardware Library or the New Project form.)_")
        out = project_path / "metadata" / "HW01_BOM.md"
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text("\n".join(lines), encoding="utf-8")
        return out

    # BOM summary table
    lines.append("## Device List")
    lines.append("")
    lines.append("| # | Device ID | Vendor | Model | Quantity | Category | Source |")
    lines.append("|---|-----------|--------|-------|----------|----------|--------|")

    for idx, dev in enumerate(selected_devices, start=1):
        device_id = dev.get("device_id", "?")
        info = catalog.get(device_id, {})
        # Source: project-local > factory
        proj_hw = project_path / "_hardware"
        src = "factory"
        if proj_hw.exists():
            for md in proj_hw.rglob("*.md"):
                if not md.name.startswith("_"):
                    try:
                        for line in md.read_text(encoding="utf-8").splitlines():
                            if line.strip().startswith("device_id:") and \
                               line.split(":", 1)[1].strip().strip('"') == device_id:
                                src = "project_local"
                    except Exception as exc:
                        # O-1 fix: file-read failure is logged; src defaults to factory.
                        _warn(f"generate_hw01_md: cannot read {md}: {exc}")
        lines.append(
            f"| {idx} | `{device_id}` | {info.get('vendor','—')} "
            f"| {info.get('model','—')} | {dev.get('quantity', 1)} "
            f"| {info.get('category','—')} | {src} |"
        )
    lines.append("")

    # Detail section per device (critical for the AI brief)
    lines.append("---")
    lines.append("")
    lines.append("## Device Technical Details (For AI Code Generation)")
    lines.append("")
    lines.append("> This section is used during SCL code generation. The AI reads this info; it does not invent it.")
    lines.append("")

    for dev in selected_devices:
        device_id = dev.get("device_id", "?")
        quantity  = dev.get("quantity", 1)
        notes     = dev.get("notes", "")

        md_path = find_device_md(device_id, project_path)
        info     = catalog.get(device_id, {})

        lines.append(f"### {device_id}  (Quantity: {quantity})")
        if notes:
            lines.append(f"**Project note:** {notes}")
        lines.append("")

        if md_path and md_path.exists():
            try:
                md_content = md_path.read_text(encoding="utf-8")
                # Do not embed the whole content — only the critical sections
                relevant_sections = _extract_relevant_sections(md_content, include_scl_snippets)
                lines.append(f"_Source: `{md_path.relative_to(FACTORY_ROOT)}`_")
                lines.append("")
                lines.extend(relevant_sections)
            except Exception as e:
                lines.append(f"_(MD could not be read: {e})_")
        else:
            lines.append(
                f"**Library MD not found: `{device_id}`**\n\n"
                "To add this device to the library:\n"
                "1. GUI -> Hardware Library -> Extraction Prompt\n"
                "2. Take the prompt, give the device PDF to the AI, paste the output into 'Add Device'"
            )
        lines.append("")
        lines.append("---")
        lines.append("")

    lines.append(f"*Auto-generated: AUTOMATION_FACTORY bom_manager.py — {ts}*")

    out = project_path / "metadata" / "HW01_BOM.md"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text("\n".join(lines), encoding="utf-8")
    return out


def _extract_relevant_sections(md_content: str, include_scl: bool) -> list[str]:
    """Extract the sections the AI needs from the MD content."""
    lines_out = []
    current_section = ""
    current_lines: list[str] = []
    keep_sections = {
        "## 1.": True,    # General info
        "## 2.": True,    # Communication interfaces
        "## 3.": True,    # PROFINET configuration
        "## 4.": True,    # Control words
        "## 5.": True,    # Parameters
        "## 6.": include_scl,  # SCL template
        "## 7.": False,   # Issues (long, skip)
        "## 8.": True,    # Notes
    }

    all_lines = md_content.splitlines()
    in_metadata = False
    past_metadata = False

    for line in all_lines:
        # Skip the metadata YAML block
        if not past_metadata:
            if line.strip() == "```yaml":
                in_metadata = True
                continue
            if in_metadata and line.strip() == "```":
                in_metadata = False
                past_metadata = True
                continue
            if in_metadata:
                continue

        # Section detection
        if line.startswith("## "):
            # Save the previous section
            if current_section and current_lines:
                for prefix, keep in keep_sections.items():
                    if current_section.startswith(prefix):
                        if keep:
                            lines_out.extend(current_lines)
                        break
                else:
                    pass  # Skip unknown section
            current_section = line
            current_lines = [line]
        elif current_section:
            current_lines.append(line)

    # Save the last section
    if current_section and current_lines:
        for prefix, keep in keep_sections.items():
            if current_section.startswith(prefix):
                if keep:
                    lines_out.extend(current_lines)
                break

    return lines_out


# -- CLI ----------------------------------------------------------------------

def main():
    import argparse
    p = argparse.ArgumentParser(description="Hardware BOM Manager")
    p.add_argument("--generate", metavar="PROJECT_PATH",
                   help="Generate BOM xlsx (_input/hardware_BOM.xlsx)")
    p.add_argument("--hw01", metavar="PROJECT_PATH",
                   help="Generate HW01_BOM.md (metadata/)")
    p.add_argument("--no-scl", action="store_true",
                   help="Do not include SCL templates in HW01")
    args = p.parse_args()

    project_path = Path(args.generate or args.hw01 or ".")
    state_file = project_path / "PROJECT_STATE.json"
    if not state_file.exists():
        print(f"PROJECT_STATE.json not found: {state_file}")
        return

    state = json.loads(state_file.read_text(encoding="utf-8"))
    selected = state.get("selected_devices", [])
    print(f"Selected devices: {len(selected)}")

    if args.generate:
        ok = generate_bom_xlsx(project_path, selected)
        if ok:
            print(f"BOM xlsx: {project_path / '_input' / 'hardware_BOM.xlsx'}")
        else:
            print("Could not generate BOM xlsx")

    if args.hw01:
        out = generate_hw01_md(project_path, selected, include_scl_snippets=not args.no_scl)
        print(f"HW01: {out}")


if __name__ == "__main__":
    main()

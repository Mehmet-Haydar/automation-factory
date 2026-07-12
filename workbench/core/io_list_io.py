"""
io_list_io.py — Read/write the RD01 IO list in both Markdown (canonical) and
XLSX (export) form. Delegates MD parsing to script_md_to_xlsx.parse_md_template
to keep one source of truth for table extraction.

The canonical 15-column header (from RD01_IO_List.md):

  Tag | Address | Type | Direction | Equipment | Description |
  NormalState | EngUnit | RangeMin | RangeMax | SafetyRelated |
  SourceModule | OldTag | Notes | Status
"""

from __future__ import annotations

import sys
from dataclasses import dataclass, field, asdict, fields
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional


FACTORY_ROOT = Path(__file__).resolve().parent.parent.parent
SCRIPTS_DIR = FACTORY_ROOT / "05_SCRIPTS"
# script_md_to_xlsx lives in the dev/ tools subfolder (see 05_SCRIPTS/dev/).
# parse_md_template is the single source-of-truth MD table parser used by the
# workbench at runtime, so its real location must be on sys.path.
SCRIPTS_DEV_DIR = SCRIPTS_DIR / "dev"

CANONICAL_HEADER = [
    "Tag", "Address", "Type", "Direction", "Equipment",
    "Description", "NormalState", "EngUnit", "RangeMin", "RangeMax",
    "SafetyRelated", "SourceModule", "OldTag", "Notes", "Status",
]


@dataclass
class IORow:
    tag: str = ""
    address: str = ""
    dtype: str = ""
    direction: str = ""
    equipment: str = ""
    description: str = ""
    normal_state: str = ""
    eng_unit: str = ""
    range_min: str = ""
    range_max: str = ""
    safety_related: str = ""
    source_module: str = ""
    old_tag: str = ""
    notes: str = ""
    status: str = "Active"

    def as_cells(self) -> list[str]:
        return [
            self.tag, self.address, self.dtype, self.direction,
            self.equipment, self.description, self.normal_state,
            self.eng_unit, self.range_min, self.range_max,
            self.safety_related, self.source_module, self.old_tag,
            self.notes, self.status,
        ]

    @classmethod
    def from_cells(cls, cells: list[str]) -> "IORow":
        cells = list(cells) + [""] * (15 - len(cells))
        return cls(
            tag=cells[0], address=cells[1], dtype=cells[2],
            direction=cells[3], equipment=cells[4], description=cells[5],
            normal_state=cells[6], eng_unit=cells[7], range_min=cells[8],
            range_max=cells[9], safety_related=cells[10],
            source_module=cells[11], old_tag=cells[12],
            notes=cells[13], status=cells[14] or "Active",
        )


# ─────────────────────────────────────────────────────────────────────────────
# MD read/write
# ─────────────────────────────────────────────────────────────────────────────

def _import_md_parser():
    # script_md_to_xlsx lives under 05_SCRIPTS/dev/, so both 05_SCRIPTS and
    # 05_SCRIPTS/dev must be importable (the latter is where the module sits).
    for _p in (SCRIPTS_DIR, SCRIPTS_DEV_DIR):
        if str(_p) not in sys.path:
            sys.path.insert(0, str(_p))
    from script_md_to_xlsx import parse_md_template  # type: ignore
    return parse_md_template


def _pick_signals_table(tables: list[dict]) -> Optional[dict]:
    """Find the IO table among the tables parsed from the MD."""
    for tbl in tables:
        header = [h.lower() for h in tbl.get("header", [])]
        if header and ("tag" in header[0] or "sinyal" in header[0]):
            return tbl
    # Fallback: first table containing both "tag" and "address" headers.
    for tbl in tables:
        header = [h.lower() for h in tbl.get("header", [])]
        if "tag" in header and ("address" in header or "adres" in header):
            return tbl
    return None


def read_md(path: Path) -> tuple[list[IORow], dict]:
    """Returns (rows, frontmatter). Frontmatter is the raw YAML dict from the MD."""
    parse = _import_md_parser()
    parsed = parse(path)
    fm = parsed.get("frontmatter") or {}
    signals = _pick_signals_table(parsed.get("tables") or [])
    if signals is None:
        return [], fm

    header = signals["header"]
    idx_map = {h.strip().lower(): i for i, h in enumerate(header)}
    rows: list[IORow] = []
    for cells in signals["rows"]:
        row = IORow(
            tag=_pick(cells, idx_map, "tag"),
            address=_pick(cells, idx_map, "address", "adres"),
            dtype=_pick(cells, idx_map, "type", "tip"),
            direction=_pick(cells, idx_map, "direction", "dir", "yön", "yon"),
            equipment=_pick(cells, idx_map, "equipment", "ekipman"),
            description=_pick(cells, idx_map, "description", "açıklama", "aciklama"),
            normal_state=_pick(cells, idx_map, "normalstate", "normal_state"),
            eng_unit=_pick(cells, idx_map, "engunit", "eng_unit", "unit"),
            range_min=_pick(cells, idx_map, "rangemin", "range_min"),
            range_max=_pick(cells, idx_map, "rangemax", "range_max"),
            safety_related=_pick(cells, idx_map, "safetyrelated", "safety_related", "safety"),
            source_module=_pick(cells, idx_map, "sourcemodule", "source_module", "srcmodule"),
            old_tag=_pick(cells, idx_map, "oldtag", "old_tag"),
            notes=_pick(cells, idx_map, "notes", "notlar"),
            status=_pick(cells, idx_map, "status", "durum") or "Active",
        )
        # Skip totally empty rows
        if not any(row.as_cells()):
            continue
        rows.append(row)
    return rows, fm


def _pick(cells: list[str], idx_map: dict, *keys: str) -> str:
    for k in keys:
        if k in idx_map:
            i = idx_map[k]
            if i < len(cells):
                return (cells[i] or "").strip()
    return ""


def _strip_frontmatter(text: str) -> str:
    """Return the document body with a leading --- YAML frontmatter removed."""
    lines = text.splitlines()
    if lines and lines[0].strip() == "---":
        for i in range(1, len(lines)):
            if lines[i].strip() == "---":
                return "\n".join(lines[i + 1:])
    return text


def _extract_pre_summary(text: str) -> str:
    """Content between the frontmatter and the first Summary/Signals/#UNKNOWNS
    heading (title, intro, an embedded ## Frontmatter block). '' if none."""
    if not text:
        return ""
    lines = _strip_frontmatter(text).splitlines()
    stop = None
    for i, ln in enumerate(lines):
        s = ln.strip().lower()
        if s.startswith("#") and any(
            t in s for t in ("özet", "ozet", "summary", "sinyal", "signal", "unknown")
        ):
            stop = i
            break
    if stop is None:
        return ""
    return "\n".join(lines[:stop]).strip("\n")


def _extract_tail_after_signals(text: str) -> str:
    """Everything from the first heading after the signals table to EOF —
    i.e. #UNKNOWNS plus any later sections (critical findings, notes). This is
    what must be preserved across a grid save (C3 data-loss fix). '' if none."""
    if not text:
        return ""
    lines = text.splitlines()
    n = len(lines)
    tbl_end = None
    for i in range(n - 1):
        h = lines[i].strip()
        sep = lines[i + 1].strip()
        if (
            h.startswith("|")
            and "tag" in h.lower()
            and "-" in sep
            and set(sep) <= set("|-: ")
        ):
            j = i + 2
            while j < n and lines[j].strip().startswith("|"):
                j += 1
            tbl_end = j
            break

    start = None
    if tbl_end is not None:
        k = tbl_end
        while k < n and not lines[k].lstrip().startswith("#"):
            k += 1
        if k < n:
            start = k
    if start is None:
        # Fallback: locate an #UNKNOWNS heading directly so it is never lost.
        for i, ln in enumerate(lines):
            if ln.lstrip().startswith("#") and "unknown" in ln.lower():
                start = i
                break
    if start is None:
        return ""
    return "\n".join(lines[start:]).rstrip()


# Section headings as written by write_md. _SECTION_LABELS is overridable per
# output language (see I1); kept Turkish-default for backward compatibility.
_SECTION_LABELS = {
    "title": "# RD01 — IO List",
    "summary": "## Özet",
    "total": "Toplam sinyal",
    "safety": "Safety-related",
    "signals": "## Sinyaller",
    "unknowns": "## #UNKNOWNS (Gate 3 — insan dolduracak)",
    "unknowns_header": "| Eski Tag | Sebep |",
}

_LABELS_BY_LANG: dict[str, dict[str, str]] = {
    "TR": _SECTION_LABELS,
    "EN": {
        "title": "# RD01 — IO List",
        "summary": "## Summary",
        "total": "Total signals",
        "safety": "Safety-related",
        "signals": "## Signals",
        "unknowns": "## #UNKNOWNS (Gate 3 — human to fill)",
        "unknowns_header": "| Old Tag | Reason |",
    },
    "DE": {
        "title": "# RD01 — IO-Liste",
        "summary": "## Zusammenfassung",
        "total": "Signale gesamt",
        "safety": "Sicherheitsrelevant",
        "signals": "## Signale",
        "unknowns": "## #UNKNOWNS (Gate 3 — manuell ausfüllen)",
        "unknowns_header": "| Altes Tag | Grund |",
    },
}


def _labels_for(frontmatter: dict, override: Optional[dict]) -> dict:
    lang = (frontmatter.get("output_language") or "TR").upper()
    base = _LABELS_BY_LANG.get(lang, _SECTION_LABELS)
    return {**base, **(override or {})}


def write_md(
    path: Path,
    rows: Iterable[IORow],
    frontmatter: Optional[dict] = None,
    labels: Optional[dict] = None,
) -> None:
    """Write the canonical MD layout (frontmatter + Summary + Signals table +
    preserved tail). Frontmatter, Summary and the Signals table are regenerated;
    the #UNKNOWNS section and any later sections in an existing file are
    preserved verbatim so a grid save never silently discards them (C3)."""
    fm = dict(frontmatter or {})
    lbl = _labels_for(fm, labels)
    if "filled_at" not in fm:
        fm["filled_at"] = datetime.now().strftime("%Y-%m-%d")

    rows_list = list(rows)
    di = sum(1 for r in rows_list if r.direction.upper() == "DI")
    do = sum(1 for r in rows_list if r.direction.upper() == "DO")
    ai = sum(1 for r in rows_list if r.direction.upper() == "AI")
    ao = sum(1 for r in rows_list if r.direction.upper() == "AO")
    safety = sum(1 for r in rows_list if r.safety_related.upper() == "Y")

    # Read the existing file so non-signal content can be preserved.
    existing = ""
    try:
        if path and Path(path).is_file():
            existing = Path(path).read_text(encoding="utf-8")
    except Exception:
        existing = ""
    pre_section = _extract_pre_summary(existing)
    tail_section = _extract_tail_after_signals(existing)

    lines: list[str] = []
    lines.append("---")
    # M-A3: serialize frontmatter via yaml.safe_dump so values containing
    # ':', '#', '\n', lists, dicts, or quotes survive the round trip. The
    # old naive `f"{key}: {value}"` interpolation broke YAML the moment a
    # customer name had a colon ("Bosch: Bonnenfant") or a description
    # spanned multiple lines.
    try:
        import yaml  # type: ignore
        dumped = yaml.safe_dump(
            fm, sort_keys=False, allow_unicode=True, default_flow_style=False,
        )
        # safe_dump ends with a newline — split into lines (drop trailing
        # empty entry) so we keep our explicit '---' delimiters.
        for ln in dumped.rstrip("\n").splitlines():
            lines.append(ln)
    except Exception:
        # Fallback: best-effort key:value escaping so we still produce
        # something parseable even if PyYAML is unavailable.
        for key, value in fm.items():
            s = str(value)
            if any(ch in s for ch in (":", "#", "\n", '"')) or s != s.strip():
                s = '"' + s.replace('\\', '\\\\').replace('"', '\\"').replace("\n", "\\n") + '"'
            lines.append(f"{key}: {s}")
    lines.append("---")
    lines.append("")
    if pre_section:
        lines.append(pre_section.rstrip())
        lines.append("")
    else:
        lines.append(lbl["title"])
        lines.append("")
    lines.append(lbl["summary"])
    lines.append("")
    lines.append(f"- {lbl['total']}: {len(rows_list)}")
    lines.append(f"- DI: {di} | DO: {do} | AI: {ai} | AO: {ao}")
    lines.append(f"- {lbl['safety']}: {safety}")
    lines.append("")
    lines.append(lbl["signals"])
    lines.append("")
    lines.append("| " + " | ".join(CANONICAL_HEADER) + " |")
    lines.append("|" + "|".join("-----" for _ in CANONICAL_HEADER) + "|")
    for row in rows_list:
        cells = [c.replace("|", "\\|") for c in row.as_cells()]
        lines.append("| " + " | ".join(cells) + " |")
    lines.append("")
    if tail_section:
        lines.append(tail_section.rstrip())
        lines.append("")
    else:
        lines.append(lbl["unknowns"])
        lines.append("")
        lines.append(lbl["unknowns_header"])
        lines.append("|----------|-------|")
        lines.append("| | |")
        lines.append("")

    text = "\n".join(lines)
    _atomic_write(path, text)


# ─────────────────────────────────────────────────────────────────────────────
# XLSX read/write
# ─────────────────────────────────────────────────────────────────────────────

def read_xlsx(path: Path) -> tuple[list[IORow], dict]:
    """Read the IO list from an XLSX. Returns (rows, frontmatter={})."""
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as e:
        raise RuntimeError("openpyxl is required to read XLSX IO lists") from e

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    # Prefer a sheet named "Signals"; otherwise the first sheet.
    sheet = wb["Signals"] if "Signals" in wb.sheetnames else wb[wb.sheetnames[0]]

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], {}
    header_lower = [(c or "").strip().lower() if isinstance(c, str) else "" for c in header]
    idx_map = {h: i for i, h in enumerate(header_lower) if h}

    rows: list[IORow] = []
    for raw_row in rows_iter:
        if not any(raw_row):
            continue
        cells = [("" if c is None else str(c)) for c in raw_row]
        row = IORow(
            tag=_pick(cells, idx_map, "tag"),
            address=_pick(cells, idx_map, "address", "adres"),
            dtype=_pick(cells, idx_map, "type", "tip"),
            direction=_pick(cells, idx_map, "direction", "dir", "yön", "yon"),
            equipment=_pick(cells, idx_map, "equipment", "ekipman"),
            description=_pick(cells, idx_map, "description", "açıklama", "aciklama"),
            normal_state=_pick(cells, idx_map, "normalstate", "normal_state"),
            eng_unit=_pick(cells, idx_map, "engunit", "eng_unit", "unit"),
            range_min=_pick(cells, idx_map, "rangemin", "range_min"),
            range_max=_pick(cells, idx_map, "rangemax", "range_max"),
            safety_related=_pick(cells, idx_map, "safetyrelated", "safety_related", "safety"),
            source_module=_pick(cells, idx_map, "sourcemodule", "source_module", "srcmodule"),
            old_tag=_pick(cells, idx_map, "oldtag", "old_tag"),
            notes=_pick(cells, idx_map, "notes", "notlar"),
            status=_pick(cells, idx_map, "status", "durum") or "Active",
        )
        if not any(row.as_cells()):
            continue
        rows.append(row)
    return rows, {}


def write_xlsx(path: Path, rows: Iterable[IORow], frontmatter: Optional[dict] = None) -> None:
    """Write the IO list to an XLSX with sheet name 'Signals'."""
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Font, PatternFill  # type: ignore
    except ImportError as e:
        raise RuntimeError("openpyxl is required to write XLSX IO lists") from e

    wb = Workbook()
    ws = wb.active
    ws.title = "Signals"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
    ws.append(CANONICAL_HEADER)
    for col_idx, _ in enumerate(CANONICAL_HEADER, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    for row in rows:
        ws.append(row.as_cells())

    # Best-effort column widths
    widths = [20, 14, 8, 10, 14, 30, 10, 8, 10, 10, 10, 14, 16, 20, 8]
    for col_idx, w in enumerate(widths, start=1):
        try:
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
        except Exception:
            pass

    wb.save(str(path))


def sync(md_path: Path, rows: Iterable[IORow], frontmatter: Optional[dict] = None) -> Path:
    """Write canonical MD + sibling XLSX. Returns the MD path."""
    rows_list = list(rows)
    write_md(md_path, rows_list, frontmatter=frontmatter)
    xlsx_path = md_path.with_suffix(".xlsx")
    try:
        write_xlsx(xlsx_path, rows_list, frontmatter=frontmatter)
    except RuntimeError:
        pass
    return md_path


def _atomic_write(path: Path, text: str) -> None:
    """Write text via temp + replace so an interrupted save can't corrupt the file."""
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(text, encoding="utf-8")
    tmp.replace(path)
